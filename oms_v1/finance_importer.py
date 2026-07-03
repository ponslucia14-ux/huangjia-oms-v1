from __future__ import annotations

import csv
import json
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from .live_connector import DEFAULT_LIVE_ROOT
from .operating_center_source import OPERATING_CENTER_PEOPLE, OPERATING_CENTER_VERSION
from .schemas import new_id, now_iso


FINANCE_IMPORT_SCHEMA_VERSION = "oms.v1.finance_import_stream"
FINANCE_METADATA_KEYS = {"__source_sheet", "__row_number"}
FINANCE_XLSX_SCAN_MAX_COLUMNS = 256
FINANCE_HEADER_SCAN_ROWS = 30
FINANCE_RECOGNIZED_HEADERS = {
    "序号",
    "日期",
    "签约日期",
    "合同编号",
    "宝妈",
    "宝妈姓名",
    "姓名",
    "客户",
    "收入",
    "收入项目",
    "收入金额",
    "支出",
    "支出项目",
    "支出金额",
    "余额",
    "备注",
    "全款费用",
    "已交房款",
    "定金",
    "已交尾款",
    "未交尾款",
    "扣手续费",
    "实际到账金额",
    "价格",
    "退款",
    "退费",
    "房款",
    "原定房款",
    "服务日房款",
    "服务金额",
    "年服务金额",
    "提成",
    "工资",
    "应发",
    "销售",
    "顾问",
    "管家",
    "照护师",
}
FINANCE_MEANINGFUL_VALUE_EXCLUDE_KEYS = {"序号"}
FINANCE_LEGACY_POLICY = {
    "Excel": "financial_input_source",
    "OMS": "financial_work_entry",
    "飞书": "external_sync_channel",
}


FINANCE_SOURCE_CONFIG = {
    "checkin_registration": {
        "source_name": "入住登记表",
        "workspace_key": "liujie",
        "finance_category": "income",
        "finance_flow": "财务流",
        "structure_layer": "business_layer",
        "structure_unit": "财务流",
        "daily_process": "入住财务确认",
        "action_type": "finance_checkin_registration_task",
        "settlement_type": "income_settlement",
    },
    "finance_daily": {
        "source_name": "财务日报表",
        "workspace_key": "liujie",
        "finance_category": "income_expense",
        "finance_flow": "财务流",
        "structure_layer": "business_layer",
        "structure_unit": "财务流",
        "daily_process": "财务日报复核",
        "action_type": "finance_daily_report_task",
        "settlement_type": "daily_reconciliation",
    },
    "bank_cash_journal": {
        "source_name": "银行现金日记账",
        "workspace_key": "liujie",
        "finance_category": "income_expense",
        "finance_flow": "财务流",
        "structure_layer": "business_layer",
        "structure_unit": "财务流",
        "daily_process": "银行现金对账",
        "action_type": "finance_bank_cash_journal_task",
        "settlement_type": "cash_bank_reconciliation",
    },
    "real_income": {
        "source_name": "实入账",
        "workspace_key": "liujie",
        "finance_category": "income",
        "finance_flow": "财务流",
        "structure_layer": "business_layer",
        "structure_unit": "财务流",
        "daily_process": "实入账确认",
        "action_type": "finance_real_income_task",
        "settlement_type": "income_settlement",
    },
    "service_refund": {
        "source_name": "服务金额及退费",
        "workspace_key": "liujie",
        "finance_category": "income_refund",
        "finance_flow": "财务流",
        "structure_layer": "business_layer",
        "structure_unit": "财务流",
        "daily_process": "服务金额与退费核算",
        "action_type": "finance_service_refund_task",
        "settlement_type": "service_amount_refund",
    },
    "sales_commission": {
        "source_name": "销售提成明细",
        "workspace_key": "huanhuan",
        "finance_category": "commission",
        "finance_flow": "销售结算流",
        "structure_layer": "sales_flow",
        "structure_unit": "销售结算流",
        "daily_process": "销售提成结算",
        "action_type": "finance_sales_commission_task",
        "settlement_type": "commission_settlement",
    },
    "care_wage": {
        "source_name": "照护师拆分工资表",
        "workspace_key": "boss",
        "finance_category": "wage",
        "finance_flow": "成本核算",
        "structure_layer": "system_capability_layer",
        "structure_unit": "成本核算",
        "daily_process": "照护师工资成本核算",
        "action_type": "finance_care_wage_task",
        "settlement_type": "wage_cost_accounting",
    },
    "sales_detail": {
        "source_name": "销售明细表",
        "workspace_key": "huanhuan",
        "finance_category": "sales_income",
        "finance_flow": "销售结算流",
        "structure_layer": "sales_flow",
        "structure_unit": "销售结算流",
        "daily_process": "销售明细结算",
        "action_type": "finance_sales_detail_task",
        "settlement_type": "sales_income_settlement",
    },
}


class FinanceDataImporter:
    """Convert finance Excel rows into OMS financial work items and settlement records."""

    def __init__(self, live_root: str | Path | None = None, operating_root: str | Path | None = None):
        self.live_root = Path(live_root or os.getenv("OMS_LIVE_ROOT") or DEFAULT_LIVE_ROOT)
        self.operating_root = Path(operating_root or self.live_root / "operational_core")
        self.converted_root = self.live_root / "finance_xlsx_work"

    def import_sources(self, **source_paths: str | Path | None) -> dict[str, Any]:
        filtered_paths = {key: value for key, value in source_paths.items() if value}
        records: list[dict[str, Any]] = []
        errors: list[dict[str, str]] = []
        for source_type, raw_path in filtered_paths.items():
            if source_type not in FINANCE_SOURCE_CONFIG:
                errors.append({"source_type": source_type, "path": str(raw_path), "error": "unsupported finance source type"})
                continue
            try:
                rows = self._read_table(Path(raw_path), source_type)
            except Exception as exc:
                errors.append({"source_type": source_type, "path": str(raw_path), "error": str(exc)})
                continue
            records.extend(self._record(source_type, row, index + 1, Path(raw_path)) for index, row in enumerate(rows))

        financial_events = [self._financial_event(record) for record in records]
        settlement_records = [self._settlement_record(record, event) for record, event in zip(records, financial_events)]
        work_items = [
            self._work_item(record, event, settlement)
            for record, event, settlement in zip(records, financial_events, settlement_records)
        ]
        pending = [
            self._pending_outbox(record, event, settlement, work_item)
            for record, event, settlement, work_item in zip(records, financial_events, settlement_records, work_items)
        ]
        self._persist("finance_work_items.jsonl", work_items, self.operating_root)
        self._persist("financial_events.jsonl", financial_events, self.live_root / "finance")
        self._persist("settlement_records.jsonl", settlement_records, self.live_root / "finance")
        self._persist("Finance_OMS导入.jsonl", pending, self.live_root / "pending_outbox")
        return {
            "schema_version": FINANCE_IMPORT_SCHEMA_VERSION,
            "source_of_truth": "Finance Excel",
            "people_model_source": OPERATING_CENTER_VERSION,
            "input_sources": {key: str(path) for key, path in filtered_paths.items()},
            "record_count": len(records),
            "financial_event_count": len(financial_events),
            "settlement_record_count": len(settlement_records),
            "work_item_count": len(work_items),
            "pending_outbox_count": len(pending),
            "records": records,
            "financial_events": financial_events,
            "settlement_records": settlement_records,
            "work_items": work_items,
            "pending_outbox": pending,
            "errors": errors,
            "audit": {
                "created_at": now_iso(),
                "live_root": str(self.live_root),
                "operating_root": str(self.operating_root),
                "principle": "财务Excel进入 user_id -> workspace -> role 的财务工作流、事件和结算体系。",
            },
        }

    def _read_table(self, path: Path, source_type: str) -> list[dict[str, Any]]:
        if not path.exists():
            raise FileNotFoundError(f"Finance source not found: {path}")
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                return [
                    {**self._clean_row(row), "__row_number": index + 2, "__source_sheet": ""}
                    for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter))
                ]
        if suffix == ".xls":
            path = self._convert_xls_to_xlsx(path)
            suffix = ".xlsx"
        if suffix in {".xlsx", ".xlsm"}:
            rows = self._read_xlsx(path)
            if not rows and source_type == "real_income":
                return self._read_real_income_crosstab(path)
            return rows
        raise ValueError(f"Unsupported finance source type: {suffix}")

    def _convert_xls_to_xlsx(self, path: Path) -> Path:
        self.converted_root.mkdir(parents=True, exist_ok=True)
        target = self.converted_root / f"{path.stem}.xlsx"
        script = r"""
$excel = $null
try {
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $workbook = $excel.Workbooks.Open($env:OMS_XLS_SOURCE, 0, $true)
  $workbook.SaveAs($env:OMS_XLS_TARGET, 51)
  $workbook.Close($false)
} finally {
  if ($excel) {
    $excel.Quit() | Out-Null
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
  }
}
"""
        env = os.environ.copy()
        env["OMS_XLS_SOURCE"] = str(path)
        env["OMS_XLS_TARGET"] = str(target)
        subprocess.run(["powershell", "-NoProfile", "-Command", script], check=True, env=env, capture_output=True, text=True)
        return target

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read finance .xlsx sources in this runtime") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                result.extend(self._read_xlsx_sheet(sheet))
            return result
        finally:
            workbook.close()

    def _read_real_income_crosstab(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read finance .xlsx sources in this runtime") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                month_row_index = self._real_income_month_row(rows)
                if month_row_index is None:
                    continue
                month_row = rows[month_row_index]
                months = {
                    index: str(value).strip()
                    for index, value in enumerate(month_row)
                    if value not in {"", None} and "月" in str(value)
                }
                for row_number, values in enumerate(rows[month_row_index + 1 :], start=month_row_index + 2):
                    label = str(values[0] or "").strip() if values else ""
                    if not label:
                        continue
                    for col_index, month in months.items():
                        value = values[col_index] if col_index < len(values) else ""
                        if value in {"", None}:
                            continue
                        result.append(
                            {
                                "日期": month,
                                "收入项目": label,
                                "收入金额": value,
                                "__source_sheet": sheet.title,
                                "__row_number": row_number,
                            }
                        )
            return result
        finally:
            workbook.close()

    def _real_income_month_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        for index, values in enumerate(rows[:FINANCE_HEADER_SCAN_ROWS]):
            month_count = sum(1 for value in values if value not in {"", None} and "月" in str(value))
            if month_count >= 2:
                return index
        return None

    def _read_xlsx_sheet(self, sheet: Any) -> list[dict[str, Any]]:
        max_col = min(sheet.max_column or 1, FINANCE_XLSX_SCAN_MAX_COLUMNS)
        scan_rows = list(
            sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row or 1, FINANCE_HEADER_SCAN_ROWS),
                max_col=max_col,
                values_only=True,
            )
        )
        header_position = self._detect_header_row(scan_rows)
        if header_position is None:
            return []
        headers = self._headers(scan_rows[header_position - 1])
        result: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_position + 1, max_col=max_col, values_only=True),
            start=header_position + 1,
        ):
            row = {
                headers[index]: values[index] if index < len(values) else ""
                for index in range(len(headers))
                if headers[index]
            }
            cleaned = self._clean_row(row)
            if not self._is_finance_data_row(cleaned):
                continue
            cleaned["__source_sheet"] = sheet.title
            cleaned["__row_number"] = row_number
            result.append(cleaned)
        return result

    def _detect_header_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        best: tuple[int, int] | None = None
        for index, values in enumerate(rows, start=1):
            labels = [str(value).strip() for value in values if value not in {"", None}]
            hits = sum(1 for label in labels if label in FINANCE_RECOGNIZED_HEADERS)
            if hits < 2:
                continue
            score = hits * 10 + len(labels)
            if best is None or score > best[1]:
                best = (index, score)
        return best[0] if best else None

    def _headers(self, values: tuple[Any, ...]) -> list[str]:
        headers: list[str] = []
        seen: dict[str, int] = {}
        for value in values:
            header = str(value or "").strip()
            if not header:
                headers.append("")
                continue
            count = seen.get(header, 0)
            seen[header] = count + 1
            headers.append(header if count == 0 else f"{header}_{count + 1}")
        return headers

    def _is_finance_data_row(self, row: dict[str, Any]) -> bool:
        values = [value for key, value in row.items() if key not in FINANCE_METADATA_KEYS and value not in {"", None}]
        if not values:
            return False
        meaningful_values = [
            value
            for key, value in row.items()
            if key not in FINANCE_METADATA_KEYS
            and key not in FINANCE_MEANINGFUL_VALUE_EXCLUDE_KEYS
            and value not in {"", None}
            and str(value).strip()
        ]
        if not meaningful_values:
            return False
        header_like_hits = sum(1 for value in values if str(value).strip() in FINANCE_RECOGNIZED_HEADERS)
        return header_like_hits < max(2, len(values) // 2)

    def _record(self, source_type: str, row: dict[str, Any], row_number: int, path: Path) -> dict[str, Any]:
        config = FINANCE_SOURCE_CONFIG[source_type]
        person = OPERATING_CENTER_PEOPLE[config["workspace_key"]]
        user_id = os.getenv(person["feishu_env"], "").strip()
        business_row = {key: value for key, value in row.items() if key not in FINANCE_METADATA_KEYS}
        normalized = self._normalized_fields(source_type, business_row)
        return {
            "record_id": new_id("fin"),
            "source_type": source_type,
            "source_name": config["source_name"],
            "source_file": str(path),
            "source_sheet": str(row.get("__source_sheet", "")),
            "row_number": int(row.get("__row_number") or row_number),
            "raw_row": business_row,
            "normalized": normalized,
            "finance_mapping": {
                "category": config["finance_category"],
                "flow": config["finance_flow"],
                "structure_layer": config["structure_layer"],
                "structure_unit": config["structure_unit"],
            },
            "assignment": {
                "user_id": user_id,
                "user_id_status": "mapped" if user_id else "unresolved_user_id",
                "workspace_key": config["workspace_key"],
                "workspace": person["title"],
                "role": self._business_role(config["workspace_key"]),
                "name": person["name"],
            },
            "structure": self._structure(config),
        }

    def _financial_event(self, record: dict[str, Any]) -> dict[str, Any]:
        return {
            "schema_version": "oms.v1.financial_event",
            "financial_event_id": new_id("fevt"),
            "record_id": record["record_id"],
            "event_type": f"finance_{record['finance_mapping']['category']}",
            "source_of_truth": "Finance Excel",
            "source_type": record["source_type"],
            "source_sheet": record["source_sheet"],
            "row_number": record["row_number"],
            "customer_name": record["normalized"]["customer_name"],
            "amount": record["normalized"]["amount"],
            "income_amount": record["normalized"]["income_amount"],
            "expense_amount": record["normalized"]["expense_amount"],
            "settlement_subject": record["normalized"]["settlement_subject"],
            "occurred_at": record["normalized"]["date"],
            "structure": record["structure"],
            "assignment": record["assignment"],
            "created_at": now_iso(),
        }

    def _settlement_record(self, record: dict[str, Any], event: dict[str, Any]) -> dict[str, Any]:
        config = FINANCE_SOURCE_CONFIG[record["source_type"]]
        return {
            "schema_version": "oms.v1.settlement_record",
            "settlement_id": new_id("set"),
            "financial_event_id": event["financial_event_id"],
            "record_id": record["record_id"],
            "settlement_type": config["settlement_type"],
            "status": "pending_confirmation",
            "amount": record["normalized"]["amount"],
            "income_amount": record["normalized"]["income_amount"],
            "expense_amount": record["normalized"]["expense_amount"],
            "workspace": record["assignment"]["workspace"],
            "role": record["assignment"]["role"],
            "user_id": record["assignment"]["user_id"],
            "user_id_status": record["assignment"]["user_id_status"],
            "source_file": record["source_file"],
            "source_sheet": record["source_sheet"],
            "row_number": record["row_number"],
            "created_at": now_iso(),
        }

    def _work_item(self, record: dict[str, Any], event: dict[str, Any], settlement: dict[str, Any]) -> dict[str, Any]:
        config = FINANCE_SOURCE_CONFIG[record["source_type"]]
        assignment = record["assignment"]
        return {
            "schema_version": "oms.v1.operational_work_item",
            "work_item_id": new_id("op"),
            "role": assignment["role"],
            "workspace": assignment["workspace"],
            "daily_process": config["daily_process"],
            "primary_entry": "OMS",
            "legacy_policy": FINANCE_LEGACY_POLICY,
            "action_id": record["record_id"],
            "action_type": config["action_type"],
            "status": "ready_with_pending_sync" if assignment["user_id"] else "attention_required",
            "confirmation_required": not bool(assignment["user_id"]),
            "next_operator_action": self._next_action(record),
            "source_sync_targets": [config["source_name"], "OMS financial_events", "OMS settlement_records", "OMS pending_outbox"],
            "financial_event_id": event["financial_event_id"],
            "settlement_id": settlement["settlement_id"],
            "finance_record": record,
        }

    def _pending_outbox(
        self,
        record: dict[str, Any],
        event: dict[str, Any],
        settlement: dict[str, Any],
        work_item: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "created_at": now_iso(),
            "mode": "Finance_Pending_Mode",
            "source_of_truth": "Finance Excel",
            "people_model_source": OPERATING_CENTER_VERSION,
            "status": "pending",
            "reason": "Finance Excel row converted to OMS financial_event, settlement_record, and work_item; external sync waits for live connector availability.",
            "record": record,
            "financial_event": event,
            "settlement_record": settlement,
            "work_item": work_item,
        }

    def _structure(self, config: dict[str, Any]) -> dict[str, Any]:
        layer = config["structure_layer"]
        support_layer = []
        system_layer = ["成本核算", "经营指标中心"]
        if config["finance_category"] in {"income_expense", "income_refund"}:
            support_layer.append("成本流")
        if config["finance_category"] == "wage":
            system_layer = ["成本核算"]
        if config["finance_category"] == "commission":
            system_layer = ["销售结算", "经营指标中心"]
        return {
            "layer": layer,
            "business_layer": ["财务流"] if layer == "business_layer" else [],
            "support_layer": support_layer,
            "system_capability_layer": system_layer,
            "sales_flow": ["销售结算流"] if layer == "sales_flow" else [],
        }

    def _normalized_fields(self, source_type: str, row: dict[str, Any]) -> dict[str, Any]:
        income = self._first(row, ["收入", "收入金额", "已交房款", "已交尾款", "实际到账金额", "全款费用", "服务日房款", "年服务金额"])
        expense = self._first(row, ["支出", "支出金额", "退款", "退费", "扣手续费", "照护师跟家工资", "工资", "应发", "提成"])
        amount = self._first(
            row,
            [
                "金额",
                "价格",
                "全款费用",
                "实际到账金额",
                "收入金额",
                "支出金额",
                "服务日房款",
                "年服务金额",
                "提成",
                "应发",
                "工资",
                "退款",
                "退费",
            ],
        )
        return {
            "business_type": source_type,
            "date": self._first(row, ["日期", "签约日期", "入住日期", "出馆日期", "预产期"]),
            "customer_name": self._first(row, ["宝妈", "宝妈姓名", "姓名", "客户"]),
            "contract_no": self._first(row, ["合同编号", "合同号", "订单号"]),
            "settlement_subject": self._first(row, ["收入项目", "支出项目", "套系", "套餐", "服务", "备注"]),
            "amount": str(amount) if amount != "" else "",
            "income_amount": str(income) if income != "" else "",
            "expense_amount": str(expense) if expense != "" else "",
            "sales_owner": self._first(row, ["销售", "顾问"]),
            "note": self._first(row, ["备注", "送餐地址"]),
        }

    def _first(self, row: dict[str, Any], names: list[str]) -> Any:
        for name in names:
            if name in row and row[name] not in {"", None}:
                return row[name]
        return ""

    def _business_role(self, workspace_key: str) -> str:
        if workspace_key == "huanhuan":
            return "销售"
        if workspace_key == "boss":
            return "管理"
        return "财务"

    def _next_action(self, record: dict[str, Any]) -> str:
        assignment = record["assignment"]
        if assignment["user_id"]:
            return f"进入 {assignment['workspace']}，由 {assignment['name']} 处理财务 Excel 来源事项。"
        return f"{assignment['workspace']} 缺少真实飞书 user_id，财务事项保留在 pending_outbox。"

    def _clean_row(self, row: dict[str, Any]) -> dict[str, Any]:
        return {str(key).strip(): "" if value is None else value for key, value in row.items() if str(key).strip()}

    def _persist(self, file_name: str, rows: list[dict[str, Any]], root: Path) -> None:
        root.mkdir(parents=True, exist_ok=True)
        path = root / file_name
        with path.open("a", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
