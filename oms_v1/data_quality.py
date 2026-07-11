from __future__ import annotations

import hashlib
import json
import math
import re
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from fnmatch import fnmatch
from pathlib import Path
from typing import Any, Iterable

from .audit_log import AuditEngine
from .event_bus import EventBus, OMSEvent
from .master_data import OMSMasterData
from .schemas import new_id, now_iso


DATA_QUALITY_SCHEMA_VERSION = "oms.v1.data_quality"
DEFAULT_DATA_QUALITY_REPORT_ROOT = Path(__file__).resolve().parents[1] / "live_runtime" / "data_quality_reports"

CURRENT_PRODUCTION = "CURRENT_PRODUCTION"
HISTORICAL = "HISTORICAL"
SUMMARY = "SUMMARY"
AUXILIARY_CALCULATION = "AUXILIARY_CALCULATION"
NOTES = "NOTES"
UNCONFIRMED = "UNCONFIRMED"

ADMITTED_CURRENT = "ADMITTED_CURRENT"
ADMITTED_HISTORICAL = "ADMITTED_HISTORICAL"
EXCLUDED_SUMMARY = "EXCLUDED_SUMMARY"
EXCLUDED_AUXILIARY = "EXCLUDED_AUXILIARY"
EXCLUDED_NOTES = "EXCLUDED_NOTES"
REVIEW_REQUIRED = "REVIEW_REQUIRED"
QUARANTINED = "QUARANTINED"
REJECTED = "REJECTED"

NEW = "NEW"
CHANGED = "CHANGED"
UNCHANGED = "UNCHANGED"
MISSING = "MISSING"
CONFLICT = "CONFLICT"

QUALITY_ADMISSIBLE = "ADMISSIBLE"
QUALITY_PARTIAL = "PARTIALLY_ADMISSIBLE"
QUALITY_REVIEW = "REVIEW_REQUIRED"
QUALITY_REJECTED = "REJECTED"

HEALTH_PASS = "PASS"
HEALTH_WARNING = "WARNING"
HEALTH_FAIL = "FAIL"

USAGE_CLASSES = {
    CURRENT_PRODUCTION,
    HISTORICAL,
    SUMMARY,
    AUXILIARY_CALCULATION,
    NOTES,
    UNCONFIRMED,
}

EXCLUSION_BY_USAGE = {
    SUMMARY: EXCLUDED_SUMMARY,
    AUXILIARY_CALCULATION: EXCLUDED_AUXILIARY,
    NOTES: EXCLUDED_NOTES,
    UNCONFIRMED: REVIEW_REQUIRED,
}

TERMINAL_STATUSES = {
    "completed",
    "cancelled",
    "canceled",
    "closed",
    "archived",
    "refunded",
    "lost",
    "inactive",
    "已完成",
    "已结束",
    "结束",
    "退款",
    "已退款",
    "流失",
    "出馆",
    "已出馆",
    "作废",
}


def _normalized_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-—–:：/\\()（）\[\]【】]+", "", text)


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=_json_value)


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_value(value: Any) -> str:
    return _sha256_bytes(_canonical_json(value).encode("utf-8"))


def _ratio(value: float) -> float:
    if not 0 <= value <= 1:
        raise ValueError("ratio must be between 0 and 1")
    return float(value)


def _parse_date_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m", "%m-%d", "%m-%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(normalized, fmt)
        except ValueError:
            continue
        if fmt.startswith("%m"):
            parsed = parsed.replace(year=datetime.now().year)
        return parsed
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _column_letter(index: int) -> str:
    if index < 1:
        raise ValueError("column index must be positive")
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


@dataclass(frozen=True)
class DataQualityPolicy:
    data_domain: str
    source_version: str
    adapter_id: str
    mapping_version: str
    required_fields: tuple[str, ...]
    business_key_fields: tuple[str, ...]
    field_aliases: dict[str, tuple[str, ...]] = field(default_factory=dict)
    field_types: dict[str, str] = field(default_factory=dict)
    status_field: str = "status"
    effective_time_field: str = "effective_time"
    expire_time_field: str = "expire_time"
    terminal_statuses: tuple[str, ...] = tuple(TERMINAL_STATUSES)
    sheet_registry: dict[str, str] = field(default_factory=dict)
    update_frequency: str = "daily"
    owner_emp_id: str = ""

    def __post_init__(self) -> None:
        if not self.data_domain.strip():
            raise ValueError("data_domain is required")
        if not self.source_version.strip():
            raise ValueError("source_version is required")
        if not self.adapter_id.strip():
            raise ValueError("adapter_id is required")
        if not self.mapping_version.strip():
            raise ValueError("mapping_version is required")
        if not self.required_fields:
            raise ValueError("required_fields is required")
        if not self.business_key_fields:
            raise ValueError("business_key_fields is required")
        invalid_classes = set(self.sheet_registry.values()) - USAGE_CLASSES
        if invalid_classes:
            raise ValueError(f"unknown sheet usage classes: {sorted(invalid_classes)}")

    def canonical_field(self, header: str) -> str:
        normalized = _normalized_text(header)
        for canonical, aliases in self.field_aliases.items():
            candidates = {_normalized_text(canonical), *(_normalized_text(alias) for alias in aliases)}
            if normalized in candidates:
                return canonical
        return str(header or "").strip()


@dataclass(frozen=True)
class SheetProfile:
    sheet_profile_id: str
    import_id: str
    source_file: str
    source_sheet: str
    sheet_index: int
    visibility: str
    row_count: int
    column_count: int
    header_row: int | None
    fields: tuple[dict[str, Any], ...]
    time_range_start: str | None
    time_range_end: str | None
    business_purpose: str | None
    usage_class: str
    classification_confidence: float
    classification_evidence: tuple[str, ...]
    admission_status: str
    issues: tuple[dict[str, Any], ...] = ()
    formula_ratio: float = 0.0
    schema_version: str = DATA_QUALITY_SCHEMA_VERSION

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["fields"] = [dict(item) for item in self.fields]
        payload["classification_evidence"] = list(self.classification_evidence)
        payload["issues"] = [dict(item) for item in self.issues]
        return payload


@dataclass(frozen=True)
class DataHealthInput:
    completeness_ratio: float
    consistency_ratio: float
    timeliness_ratio: float
    traceability_ratio: float
    anomaly_counts: dict[str, int] = field(default_factory=dict)
    hard_fail_reasons: tuple[str, ...] = ()

    def __post_init__(self) -> None:
        _ratio(self.completeness_ratio)
        _ratio(self.consistency_ratio)
        _ratio(self.timeliness_ratio)
        _ratio(self.traceability_ratio)
        if any(count < 0 for count in self.anomaly_counts.values()):
            raise ValueError("anomaly counts cannot be negative")


class DataHealthScorer:
    WEIGHTS = {
        "completeness": 25.0,
        "consistency": 30.0,
        "timeliness": 15.0,
        "traceability": 25.0,
        "anomalies": 5.0,
    }
    ANOMALY_PENALTIES = {
        "critical": 5.0,
        "high": 2.0,
        "medium": 0.5,
        "low": 0.1,
    }

    def score(self, value: DataHealthInput) -> dict[str, Any]:
        anomaly_penalty = sum(
            self.ANOMALY_PENALTIES.get(str(severity).lower(), 0.0) * int(count)
            for severity, count in value.anomaly_counts.items()
        )
        dimensions = {
            "completeness": round(self.WEIGHTS["completeness"] * value.completeness_ratio, 2),
            "consistency": round(self.WEIGHTS["consistency"] * value.consistency_ratio, 2),
            "timeliness": round(self.WEIGHTS["timeliness"] * value.timeliness_ratio, 2),
            "traceability": round(self.WEIGHTS["traceability"] * value.traceability_ratio, 2),
            "anomalies": round(max(0.0, self.WEIGHTS["anomalies"] - anomaly_penalty), 2),
        }
        total = round(sum(dimensions.values()), 2)
        hard_fail = bool(value.hard_fail_reasons)
        if hard_fail or total < 80:
            status = HEALTH_FAIL
        elif total < 95:
            status = HEALTH_WARNING
        else:
            status = HEALTH_PASS
        return {
            "schema_version": DATA_QUALITY_SCHEMA_VERSION,
            "score": total,
            "status": status,
            "dimensions": dimensions,
            "ratios": {
                "completeness": value.completeness_ratio,
                "consistency": value.consistency_ratio,
                "timeliness": value.timeliness_ratio,
                "traceability": value.traceability_ratio,
            },
            "anomaly_counts": dict(value.anomaly_counts),
            "anomaly_penalty": round(anomaly_penalty, 2),
            "hard_fail_override": hard_fail,
            "hard_fail_reasons": list(value.hard_fail_reasons),
            "calculated_at": now_iso(),
        }

    def overall(self, domain_scores: dict[str, dict[str, Any]]) -> dict[str, Any]:
        if not domain_scores:
            raise ValueError("domain_scores is required")
        score = round(sum(float(item["score"]) for item in domain_scores.values()) / len(domain_scores), 2)
        statuses = {str(item["status"]) for item in domain_scores.values()}
        if HEALTH_FAIL in statuses:
            status = HEALTH_FAIL
        elif HEALTH_WARNING in statuses:
            status = HEALTH_WARNING
        else:
            status = HEALTH_PASS
        return {
            "schema_version": DATA_QUALITY_SCHEMA_VERSION,
            "score": score,
            "status": status,
            "domain_scores": {domain: dict(value) for domain, value in domain_scores.items()},
            "calculated_at": now_iso(),
        }


@dataclass
class _SheetData:
    profile: SheetProfile
    records: list[dict[str, Any]]


class SheetSemanticMemory:
    """Persist approved Sheet semantics without bypassing later quality checks."""

    SCHEMA_VERSION = "oms.v1.sheet_semantic_memory.v2"
    CONFIRMED = "CONFIRMED"
    TEMPORARY = "TEMPORARY"
    DEPRECATED = "DEPRECATED"
    VALID_MEMORY_STATUSES = {CONFIRMED, TEMPORARY, DEPRECATED}
    STATUS_AUTO = "AUTO_RECOGNIZED"
    STATUS_REVIEW = "REVIEW_REQUIRED"
    STATUS_MISSING = "NO_HISTORY"

    def __init__(self, path: str | Path, *, audit: AuditEngine | None = None):
        self.path = Path(path)
        self.audit = audit

    def confirm(
        self,
        *,
        source_file_pattern: str,
        source_sheet: str,
        domain: str,
        fact_type: str,
        owner: str,
        profile: dict[str, Any],
        workbook_sheets: Iterable[str],
        quality_result: str,
        confirmed_at: str | None = None,
        memory_status: str = CONFIRMED,
        actor_emp_id: str = "EMP001",
        actor_name: str = "石磊",
        reason: str = "Confirm Sheet semantic memory.",
    ) -> dict[str, Any]:
        required = {
            "source_file_pattern": source_file_pattern,
            "source_sheet": source_sheet,
            "domain": domain,
            "fact_type": fact_type,
            "owner": owner,
        }
        missing = [key for key, value in required.items() if not str(value or "").strip()]
        if missing:
            raise ValueError(f"semantic memory fields required: {', '.join(missing)}")
        if memory_status not in self.VALID_MEMORY_STATUSES - {self.DEPRECATED}:
            raise ValueError("new semantic memory status must be CONFIRMED or TEMPORARY")
        payload = self._read()
        related = self._related_entries(
            payload["entries"], source_file_pattern=source_file_pattern, source_sheet=source_sheet, domain=domain
        )
        version_number = max((self._version_number(item.get("memory_version")) for item in related), default=0) + 1
        memory_id = f"sheetmem_{_sha256_value([source_file_pattern, source_sheet, domain])[:16]}"
        entry = {
            "memory_id": memory_id,
            "memory_version": f"{memory_id}-V{version_number}",
            "memory_status": memory_status,
            "source_file_pattern": source_file_pattern,
            "source_sheet": source_sheet,
            "domain": domain,
            "fact_type": fact_type,
            "owner": owner,
            "usage_class": self._usage_class(fact_type),
            "confirmed_at": confirmed_at or now_iso(),
            "quality_result": quality_result,
            "workbook_sheets": sorted({str(item) for item in workbook_sheets}),
            "field_signature": self._field_signature(profile),
            "structure_signature": self._structure_signature(profile),
        }
        entries = list(payload["entries"])
        entries.append(entry)
        self._write({"schema_version": self.SCHEMA_VERSION, "entries": entries})
        self._audit(
            action="sheet.semantic_memory.created",
            actor_emp_id=actor_emp_id,
            actor_name=actor_name,
            reason=reason,
            result=memory_status,
            after=entry,
        )
        return dict(entry)

    def transition_status(
        self,
        memory_version: str,
        *,
        status: str,
        actor_emp_id: str,
        actor_name: str,
        reason: str,
    ) -> dict[str, Any]:
        if status not in self.VALID_MEMORY_STATUSES:
            raise ValueError(f"unknown semantic memory status: {status}")
        payload = self._read()
        matches = [item for item in payload["entries"] if item.get("memory_version") == memory_version]
        if len(matches) != 1:
            raise KeyError(memory_version)
        current = matches[0]
        if current.get("memory_status") == self.DEPRECATED:
            raise ValueError("deprecated semantic memory is immutable")
        if status == current.get("memory_status"):
            return dict(current)
        if status not in {self.CONFIRMED, self.DEPRECATED}:
            raise ValueError("memory may only transition to CONFIRMED or DEPRECATED")
        before = dict(current)
        current["memory_status"] = status
        current["status_changed_at"] = now_iso()
        self._write(payload)
        self._audit(
            action="sheet.semantic_memory.status_changed",
            actor_emp_id=actor_emp_id,
            actor_name=actor_name,
            reason=reason,
            result=status,
            before=before,
            after=current,
        )
        return dict(current)

    def resolve(
        self,
        *,
        source_file: str,
        source_sheet: str,
        domain: str,
        profile: dict[str, Any],
        workbook_sheets: Iterable[str],
        quality_result: str | None = None,
    ) -> dict[str, Any]:
        matches = [
            item for item in self._read()["entries"]
            if fnmatch(source_file, str(item.get("source_file_pattern") or ""))
            and item.get("source_sheet") == source_sheet
            and str(item.get("domain") or "").lower() == domain.lower()
            and item.get("memory_status", self.CONFIRMED) != self.DEPRECATED
        ]
        if matches:
            latest_by_rule: dict[str, dict[str, Any]] = {}
            for item in matches:
                memory_id = str(item.get("memory_id") or "")
                previous = latest_by_rule.get(memory_id)
                if previous is None or self._version_number(item.get("memory_version")) > self._version_number(previous.get("memory_version")):
                    latest_by_rule[memory_id] = item
            matches = list(latest_by_rule.values())
        if len(matches) != 1:
            return {
                "status": self.STATUS_MISSING if not matches else self.STATUS_REVIEW,
                "reasons": ["semantic_memory_not_found" if not matches else "semantic_memory_ambiguous"],
            }
        memory = matches[0]
        if memory.get("memory_status", self.CONFIRMED) == self.TEMPORARY:
            return {"status": self.STATUS_REVIEW, "reasons": ["temporary_memory_requires_review"], "memory": dict(memory)}
        reasons: list[str] = []
        previous_sheets = set(memory.get("workbook_sheets") or [])
        current_sheets = {str(item) for item in workbook_sheets}
        if current_sheets - previous_sheets:
            reasons.append("sheet_added")
        if previous_sheets - current_sheets:
            reasons.append("sheet_removed")
        if memory.get("field_signature") != self._field_signature(profile):
            reasons.append("fields_changed")
        if memory.get("structure_signature") != self._structure_signature(profile):
            reasons.append("structure_changed")
        if quality_result and self._quality_rank(quality_result) < self._quality_rank(str(memory.get("quality_result") or "")):
            reasons.append("data_quality_declined")
        if reasons:
            return {"status": self.STATUS_REVIEW, "reasons": reasons, "memory": dict(memory)}
        return {"status": self.STATUS_AUTO, "reasons": [], "memory": dict(memory)}

    def entries(self) -> list[dict[str, Any]]:
        return [dict(item) for item in self._read()["entries"]]

    def _read(self) -> dict[str, Any]:
        if not self.path.is_file():
            return {"schema_version": self.SCHEMA_VERSION, "entries": []}
        payload = json.loads(self.path.read_text(encoding="utf-8"))
        if payload.get("schema_version") != self.SCHEMA_VERSION or not isinstance(payload.get("entries"), list):
            raise ValueError("invalid Sheet semantic memory store")
        return payload

    @staticmethod
    def _related_entries(
        entries: Iterable[dict[str, Any]], *, source_file_pattern: str, source_sheet: str, domain: str
    ) -> list[dict[str, Any]]:
        return [
            item for item in entries
            if item.get("source_file_pattern") == source_file_pattern
            and item.get("source_sheet") == source_sheet
            and str(item.get("domain") or "").lower() == domain.lower()
        ]

    @staticmethod
    def _version_number(value: Any) -> int:
        match = re.search(r"-V(\d+)$", str(value or ""))
        return int(match.group(1)) if match else 0

    def _audit(
        self,
        *,
        action: str,
        actor_emp_id: str,
        actor_name: str,
        reason: str,
        result: str,
        before: dict[str, Any] | None = None,
        after: dict[str, Any] | None = None,
    ) -> None:
        if self.audit is None:
            return
        self.audit.record(
            emp_id=actor_emp_id,
            actor_name=actor_name,
            module="data_quality",
            action=action,
            reason=reason,
            result=result,
            target_type="sheet_semantic_memory",
            target_id=str((after or before or {}).get("memory_version") or ""),
            before_hash=_sha256_value(before) if before is not None else "",
            after_hash=_sha256_value(after) if after is not None else "",
            metadata={"before": before, "after": after},
        )

    def _write(self, payload: dict[str, Any]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        temporary.replace(self.path)

    @staticmethod
    def _field_signature(profile: dict[str, Any]) -> str:
        fields = profile.get("fields") or []
        normalized = sorted(
            (str(item.get("source_field") or ""), str(item.get("canonical_field") or ""))
            for item in fields if isinstance(item, dict)
        )
        return _sha256_value(normalized)

    @staticmethod
    def _structure_signature(profile: dict[str, Any]) -> str:
        return _sha256_value({
            "header_row": profile.get("header_row"),
            "column_count": profile.get("column_count"),
        })

    @staticmethod
    def _usage_class(fact_type: str) -> str:
        normalized = _normalized_text(fact_type)
        return HISTORICAL if "historical" in normalized or "历史" in normalized else CURRENT_PRODUCTION

    @staticmethod
    def _quality_rank(value: str) -> int:
        return {HEALTH_FAIL: 0, QUALITY_REJECTED: 0, HEALTH_WARNING: 1, QUALITY_REVIEW: 1, HEALTH_PASS: 2, QUALITY_ADMISSIBLE: 2}.get(value, 0)


class ExcelWorkbookInspector:
    """Read every sheet and produce structure profiles without mutating the workbook."""

    def inspect(
        self,
        path: str | Path,
        *,
        import_id: str,
        policy: DataQualityPolicy,
    ) -> tuple[str, list[_SheetData]]:
        source_path = Path(path)
        if not source_path.exists():
            raise FileNotFoundError(source_path)
        file_hash = _sha256_bytes(source_path.read_bytes())
        suffix = source_path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            sheets = self._read_openpyxl(source_path, import_id=import_id, policy=policy)
        elif suffix == ".xls":
            sheets = self._read_xlrd(source_path, import_id=import_id, policy=policy)
        else:
            raise ValueError(f"unsupported Excel format: {suffix}")
        return file_hash, sheets

    def _read_openpyxl(self, path: Path, *, import_id: str, policy: DataQualityPolicy) -> list[_SheetData]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - runtime dependency guard.
            raise RuntimeError("openpyxl is required to analyze .xlsx/.xlsm files") from exc

        formula_book = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        value_book = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            result: list[_SheetData] = []
            for index, sheet_name in enumerate(formula_book.sheetnames):
                formula_sheet = formula_book[sheet_name]
                value_sheet = value_book[sheet_name]
                formula_rows = [list(row) for row in formula_sheet.iter_rows(values_only=True)]
                value_rows = [list(row) for row in value_sheet.iter_rows(values_only=True)]
                result.append(
                    self._profile_sheet(
                        path=path,
                        import_id=import_id,
                        policy=policy,
                        sheet_name=sheet_name,
                        sheet_index=index,
                        visibility=str(formula_sheet.sheet_state or "visible").upper(),
                        value_rows=value_rows,
                        formula_rows=formula_rows,
                    )
                )
            return result
        finally:
            formula_book.close()
            value_book.close()

    def _read_xlrd(self, path: Path, *, import_id: str, policy: DataQualityPolicy) -> list[_SheetData]:
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - depends on deployment package.
            raise RuntimeError("xlrd>=2 is required to analyze legacy .xls files") from exc
        workbook = xlrd.open_workbook(path, on_demand=True)
        result: list[_SheetData] = []
        try:
            for index, sheet_name in enumerate(workbook.sheet_names()):
                sheet = workbook.sheet_by_name(sheet_name)
                rows: list[list[Any]] = []
                for row_index in range(sheet.nrows):
                    values: list[Any] = []
                    for column_index in range(sheet.ncols):
                        cell = sheet.cell(row_index, column_index)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(value, workbook.datemode)
                        values.append(value)
                    rows.append(values)
                visibility = "HIDDEN" if getattr(sheet, "visibility", 0) else "VISIBLE"
                result.append(
                    self._profile_sheet(
                        path=path,
                        import_id=import_id,
                        policy=policy,
                        sheet_name=sheet_name,
                        sheet_index=index,
                        visibility=visibility,
                        value_rows=rows,
                        formula_rows=rows,
                    )
                )
            return result
        finally:
            workbook.release_resources()

    def _profile_sheet(
        self,
        *,
        path: Path,
        import_id: str,
        policy: DataQualityPolicy,
        sheet_name: str,
        sheet_index: int,
        visibility: str,
        value_rows: list[list[Any]],
        formula_rows: list[list[Any]],
    ) -> _SheetData:
        value_rows = self._trim_rows(value_rows)
        formula_rows = self._trim_rows(formula_rows)
        header_index = self._detect_header(value_rows)
        raw_headers = value_rows[header_index] if header_index is not None and value_rows else []
        headers = self._unique_headers(raw_headers)
        records: list[dict[str, Any]] = []
        if header_index is not None:
            for row_index, row in enumerate(value_rows[header_index + 1 :], start=header_index + 2):
                values = list(row) + [None] * max(0, len(headers) - len(row))
                payload = {headers[column]: _json_value(values[column]) for column in range(len(headers)) if headers[column]}
                if any(value not in {None, ""} for value in payload.values()):
                    records.append({"source_row": row_index, "values": payload})

        all_formula_cells = [value for row in formula_rows for value in row if value not in {None, ""}]
        formula_count = sum(1 for value in all_formula_cells if isinstance(value, str) and value.startswith("="))
        formula_ratio = formula_count / len(all_formula_cells) if all_formula_cells else 0.0
        fields = tuple(
            {
                "source_field": header,
                "canonical_field": policy.canonical_field(header),
                "confidence": 1.0 if policy.canonical_field(header) != header else 0.5,
            }
            for header in headers
            if header
        )
        dates = [parsed for row in value_rows for value in row if (parsed := _parse_date_value(value))]
        usage_class, confidence, evidence = self._classify(
            sheet_name=sheet_name,
            fields=fields,
            records=records,
            formula_ratio=formula_ratio,
            policy=policy,
        )
        admission_status = {
            CURRENT_PRODUCTION: ADMITTED_CURRENT,
            HISTORICAL: ADMITTED_HISTORICAL,
        }.get(usage_class, EXCLUSION_BY_USAGE.get(usage_class, REVIEW_REQUIRED))
        issues: list[dict[str, Any]] = []
        if header_index is None:
            issues.append({"code": "header_not_detected", "severity": "high"})
        if usage_class == UNCONFIRMED:
            issues.append({"code": "sheet_usage_unconfirmed", "severity": "high"})
        profile = SheetProfile(
            sheet_profile_id=new_id("sheetprof"),
            import_id=import_id,
            source_file=path.name,
            source_sheet=sheet_name,
            sheet_index=sheet_index,
            visibility=visibility,
            row_count=len(records),
            column_count=len(fields),
            header_row=header_index + 1 if header_index is not None else None,
            fields=fields,
            time_range_start=min(dates).date().isoformat() if dates else None,
            time_range_end=max(dates).date().isoformat() if dates else None,
            business_purpose=policy.data_domain if self._recognized_field_count(fields, policy) else None,
            usage_class=usage_class,
            classification_confidence=confidence,
            classification_evidence=tuple(evidence),
            admission_status=admission_status,
            issues=tuple(issues),
            formula_ratio=round(formula_ratio, 4),
        )
        return _SheetData(profile=profile, records=records)

    @staticmethod
    def _trim_rows(rows: list[list[Any]]) -> list[list[Any]]:
        cleaned = [list(row) for row in rows]
        while cleaned and not any(value not in {None, ""} for value in cleaned[-1]):
            cleaned.pop()
        max_column = 0
        for row in cleaned:
            for index, value in enumerate(row, start=1):
                if value not in {None, ""}:
                    max_column = max(max_column, index)
        return [row[:max_column] for row in cleaned]

    @staticmethod
    def _detect_header(rows: list[list[Any]]) -> int | None:
        if not rows:
            return None
        best_index: int | None = None
        best_score = -math.inf
        for index, row in enumerate(rows[:20]):
            nonempty = [value for value in row if value not in {None, ""}]
            if not nonempty:
                continue
            text_values = [str(value).strip() for value in nonempty if isinstance(value, str) and str(value).strip()]
            unique_count = len({_normalized_text(value) for value in text_values})
            next_nonempty = 0
            for candidate in rows[index + 1 : index + 4]:
                if any(value not in {None, ""} for value in candidate):
                    next_nonempty += 1
            score = len(text_values) * 3 + unique_count + next_nonempty - max(0, len(nonempty) - len(text_values))
            if score > best_score:
                best_score = score
                best_index = index
        return best_index

    @staticmethod
    def _unique_headers(values: Iterable[Any]) -> list[str]:
        result: list[str] = []
        counts: dict[str, int] = {}
        for index, value in enumerate(values, start=1):
            base = str(value or "").strip() or f"column_{index}"
            counts[base] = counts.get(base, 0) + 1
            result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
        return result

    def _classify(
        self,
        *,
        sheet_name: str,
        fields: tuple[dict[str, Any], ...],
        records: list[dict[str, Any]],
        formula_ratio: float,
        policy: DataQualityPolicy,
    ) -> tuple[str, float, list[str]]:
        registered = policy.sheet_registry.get(sheet_name)
        if registered:
            return registered, 1.0, ["approved_sheet_registry"]

        canonical_fields = {str(item["canonical_field"]) for item in fields}
        normalized_fields = {_normalized_text(item["source_field"]) for item in fields}
        required_matches = set(policy.required_fields) & canonical_fields
        key_matches = set(policy.business_key_fields) & canonical_fields
        summary_terms = {"合计", "汇总", "总计", "summary", "total"}
        note_terms = {"备注", "说明", "口径", "notes", "readme"}

        if formula_ratio >= 0.25 and not key_matches:
            return AUXILIARY_CALCULATION, 0.9, ["high_formula_ratio", f"formula_ratio={formula_ratio:.2f}"]
        if any(any(term in field for term in summary_terms) for field in normalized_fields) and not key_matches:
            return SUMMARY, 0.9, ["summary_fields_detected"]
        if len(fields) <= 2 and not required_matches and (
            any(any(term in field for term in note_terms) for field in normalized_fields) or len(records) <= 5
        ):
            return NOTES, 0.8, ["note_like_structure"]
        if set(policy.required_fields).issubset(canonical_fields) and key_matches:
            statuses = {
                _normalized_text(self._canonicalize(record["values"], policy).get(policy.status_field))
                for record in records
                if self._canonicalize(record["values"], policy).get(policy.status_field) not in {None, ""}
            }
            terminal = {_normalized_text(item) for item in policy.terminal_statuses}
            if statuses and statuses.issubset(terminal):
                return HISTORICAL, 0.9, ["required_fields_present", "all_statuses_terminal"]
            return CURRENT_PRODUCTION, 0.85, ["required_fields_present", "business_key_present"]
        return UNCONFIRMED, 0.0, ["insufficient_business_evidence"]

    @staticmethod
    def _canonicalize(values: dict[str, Any], policy: DataQualityPolicy) -> dict[str, Any]:
        return {policy.canonical_field(field): value for field, value in values.items()}

    @staticmethod
    def _recognized_field_count(fields: tuple[dict[str, Any], ...], policy: DataQualityPolicy) -> int:
        canonical = {str(item["canonical_field"]) for item in fields}
        return len(canonical & (set(policy.required_fields) | set(policy.business_key_fields)))


class DataQualityEngine:
    """Analyze Excel data and return candidates; never writes a production Truth Source."""

    def __init__(
        self,
        *,
        audit: AuditEngine | None = None,
        event_bus: EventBus | None = None,
        master_data: OMSMasterData | None = None,
        inspector: ExcelWorkbookInspector | None = None,
        report_writer: "DataQualityReportWriter | None" = None,
        report_root: str | Path | None = None,
        semantic_memory: SheetSemanticMemory | None = None,
    ):
        self.audit = audit or AuditEngine()
        self.event_bus = event_bus or EventBus()
        self.master_data = master_data or OMSMasterData()
        self.inspector = inspector or ExcelWorkbookInspector()
        self.report_writer = report_writer or DataQualityReportWriter()
        self.report_root = Path(report_root or DEFAULT_DATA_QUALITY_REPORT_ROOT)
        self.semantic_memory = semantic_memory

    def analyze_excel(
        self,
        path: str | Path,
        *,
        policy: DataQualityPolicy,
        actor_emp_id: str,
        reason: str,
        previous_records: Iterable[dict[str, Any]] = (),
        correlation_id: str = "",
        imported_at: str | None = None,
    ) -> dict[str, Any]:
        if not reason.strip():
            raise ValueError("reason is required")
        actor = self.master_data.employee_by_emp(actor_emp_id)
        if actor not in self.master_data.active_employees():
            raise PermissionError(f"inactive EMP cannot analyze production data: {actor_emp_id}")
        import_id = new_id("dqimport")
        correlation_id = correlation_id or import_id
        imported_at = imported_at or now_iso()
        source_path = Path(path)

        request_audit = self._audit(
            action="data_quality.import.request",
            actor=actor,
            reason=reason,
            result="ANALYZING",
            target_id=import_id,
            correlation_id=correlation_id,
            metadata={"source_file": source_path.name, "data_domain": policy.data_domain},
        )
        try:
            source_file_hash, sheets = self.inspector.inspect(source_path, import_id=import_id, policy=policy)
        except Exception as exc:
            failure_audit = self._audit(
                action="data_quality.admission.rejected",
                actor=actor,
                reason=str(exc),
                result=QUALITY_REJECTED,
                target_id=import_id,
                correlation_id=correlation_id,
                metadata={"source_file": source_path.name, "error": str(exc)},
            )
            event = self._event(
                "data_quality.truth_source.rejected",
                actor=actor,
                correlation_id=correlation_id,
                subject=import_id,
                payload={"import_id": import_id, "source_file": source_path.name, "error": str(exc)},
            )
            result = {
                "schema_version": DATA_QUALITY_SCHEMA_VERSION,
                "import_id": import_id,
                "source_file": source_path.name,
                "source_file_hash": "",
                "source_version": policy.source_version,
                "imported_at": imported_at,
                "quality_status": QUALITY_REJECTED,
                "sheet_profiles": [],
                "current_records": [],
                "historical_records": [],
                "quarantine_records": [],
                "missing_records": [],
                "excluded_records": [],
                "issues": [{"code": "workbook_analysis_failed", "severity": "critical", "message": str(exc)}],
                "counts": self._counts([], [], [], [], []),
                "audit_records": [request_audit, failure_audit],
                "events": [event],
                "mutates_truth_source": False,
            }
            return self._attach_report(result)

        sheets, semantic_memory_results = self._apply_semantic_memory(source_path.name, sheets, policy)
        audits = [request_audit]
        events: list[dict[str, Any]] = []
        profiles = [sheet.profile.to_dict() for sheet in sheets]
        for profile in profiles:
            audits.append(
                self._audit(
                    action="data_quality.sheet.analyzed",
                    actor=actor,
                    reason="Analyze every workbook sheet before production admission.",
                    result=profile["usage_class"],
                    target_id=profile["sheet_profile_id"],
                    correlation_id=correlation_id,
                    metadata={
                        "import_id": import_id,
                        "source_sheet": profile["source_sheet"],
                        "row_count": profile["row_count"],
                        "column_count": profile["column_count"],
                        "usage_class": profile["usage_class"],
                    },
                )
            )

        current, historical, quarantine, excluded, issues = self._evaluate_sheets(
            sheets,
            policy=policy,
            source_file_hash=source_file_hash,
            imported_at=imported_at,
            import_id=import_id,
            previous_records=list(previous_records),
        )
        seen_ids = {str(item.get("record_id") or "") for item in [*current, *historical, *quarantine]}
        missing = self._missing_records(
            previous_records=list(previous_records),
            seen_ids=seen_ids,
            import_id=import_id,
            imported_at=imported_at,
        )
        quarantine.extend(missing)
        for record in [*current, *historical, *quarantine]:
            audits.append(
                self._audit(
                    action="data_quality.record.evaluated",
                    actor=actor,
                    reason="Evaluate record quality and production admission.",
                    result=str(record["admission_status"]),
                    target_id=str(record["record_id"]),
                    correlation_id=correlation_id,
                    metadata={
                        "import_id": import_id,
                        "source_sheet": record.get("source_sheet"),
                        "source_row": record.get("source_row"),
                        "change_type": record.get("change_type"),
                        "is_current": record.get("is_current"),
                    },
                )
            )
        if any(record.get("change_type") == CONFLICT for record in quarantine):
            events.append(
                self._event(
                    "data_quality.conflict.detected",
                    actor=actor,
                    correlation_id=correlation_id,
                    subject=import_id,
                    payload={
                        "import_id": import_id,
                        "conflict_count": sum(1 for record in quarantine if record.get("change_type") == CONFLICT),
                    },
                )
            )

        unconfirmed = [profile for profile in profiles if profile["usage_class"] == UNCONFIRMED]
        admitted_count = len(current) + len(historical)
        if admitted_count and not unconfirmed and not quarantine:
            quality_status = QUALITY_ADMISSIBLE
        elif admitted_count:
            quality_status = QUALITY_PARTIAL
        elif unconfirmed or quarantine:
            quality_status = QUALITY_REVIEW
        else:
            quality_status = QUALITY_REJECTED

        analysis_event = self._event(
            "data_quality.analysis.completed",
            actor=actor,
            correlation_id=correlation_id,
            subject=import_id,
            payload={
                "import_id": import_id,
                "source_file": source_path.name,
                "sheet_count": len(profiles),
                "quality_status": quality_status,
                "admitted_count": admitted_count,
                "quarantine_count": len(quarantine),
            },
        )
        events.append(analysis_event)
        if quality_status in {QUALITY_PARTIAL, QUALITY_REVIEW}:
            events.append(
                self._event(
                    "data_quality.review.required",
                    actor=actor,
                    correlation_id=correlation_id,
                    subject=import_id,
                    payload={"import_id": import_id, "quality_status": quality_status},
                )
            )
        if admitted_count:
            events.append(
                self._event(
                    "data_quality.truth_source.admitted",
                    actor=actor,
                    correlation_id=correlation_id,
                    subject=import_id,
                    payload={
                        "import_id": import_id,
                        "current_count": len(current),
                        "historical_count": len(historical),
                        "candidate_only": True,
                    },
                )
            )
        elif quality_status == QUALITY_REJECTED:
            events.append(
                self._event(
                    "data_quality.truth_source.rejected",
                    actor=actor,
                    correlation_id=correlation_id,
                    subject=import_id,
                    payload={"import_id": import_id, "quality_status": quality_status},
                )
            )

        completion_action = "data_quality.admission.completed" if admitted_count else "data_quality.admission.rejected"
        audits.append(
            self._audit(
                action=completion_action,
                actor=actor,
                reason="Complete data quality admission analysis.",
                result=quality_status,
                target_id=import_id,
                correlation_id=correlation_id,
                metadata={
                    "current_count": len(current),
                    "historical_count": len(historical),
                    "quarantine_count": len(quarantine),
                    "mutates_truth_source": False,
                },
            )
        )
        result = {
            "schema_version": DATA_QUALITY_SCHEMA_VERSION,
            "import_id": import_id,
            "source_file": source_path.name,
            "source_file_hash": source_file_hash,
            "source_version": policy.source_version,
            "adapter_id": policy.adapter_id,
            "mapping_version": policy.mapping_version,
            "data_domain": policy.data_domain,
            "imported_at": imported_at,
            "imported_by_emp_id": actor.emp,
            "workbook_sheet_count": len(profiles),
            "quality_status": quality_status,
            "sheet_profiles": profiles,
            "current_records": current,
            "historical_records": historical,
            "quarantine_records": quarantine,
            "missing_records": missing,
            "excluded_records": excluded,
            "issues": issues,
            "counts": self._counts(current, historical, quarantine, missing, excluded),
            "audit_records": audits,
            "events": events,
            "mutates_truth_source": False,
            "semantic_memory_results": semantic_memory_results,
        }
        return self._attach_report(result)

    def _apply_semantic_memory(
        self,
        source_file: str,
        sheets: list[_SheetData],
        policy: DataQualityPolicy,
    ) -> tuple[list[_SheetData], list[dict[str, Any]]]:
        if self.semantic_memory is None:
            return sheets, []
        sheet_names = [sheet.profile.source_sheet for sheet in sheets]
        resolved_sheets: list[_SheetData] = []
        results: list[dict[str, Any]] = []
        for sheet in sheets:
            resolution = self.semantic_memory.resolve(
                source_file=source_file,
                source_sheet=sheet.profile.source_sheet,
                domain=policy.data_domain,
                profile=sheet.profile.to_dict(),
                workbook_sheets=sheet_names,
            )
            results.append({"source_sheet": sheet.profile.source_sheet, **resolution})
            if resolution["status"] == SheetSemanticMemory.STATUS_AUTO:
                usage_class = str(resolution["memory"]["usage_class"])
                admission_status = ADMITTED_HISTORICAL if usage_class == HISTORICAL else ADMITTED_CURRENT
                profile = replace(
                    sheet.profile,
                    usage_class=usage_class,
                    admission_status=admission_status,
                    classification_confidence=1.0,
                    classification_evidence=tuple([*sheet.profile.classification_evidence, "sheet_semantic_memory"]),
                )
                resolved_sheets.append(_SheetData(profile=profile, records=sheet.records))
                continue
            if resolution["status"] == SheetSemanticMemory.STATUS_REVIEW:
                profile = replace(
                    sheet.profile,
                    usage_class=UNCONFIRMED,
                    admission_status=REVIEW_REQUIRED,
                    issues=tuple([*sheet.profile.issues, {"code": "sheet_semantic_memory_review_required", "severity": "high", "reasons": resolution["reasons"]}]),
                )
                resolved_sheets.append(_SheetData(profile=profile, records=sheet.records))
                continue
            resolved_sheets.append(sheet)
        return resolved_sheets, results

    def _evaluate_sheets(
        self,
        sheets: list[_SheetData],
        *,
        policy: DataQualityPolicy,
        source_file_hash: str,
        imported_at: str,
        import_id: str,
        previous_records: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        previous_by_id = {str(item.get("record_id") or ""): item for item in previous_records if item.get("record_id")}
        current: list[dict[str, Any]] = []
        historical: list[dict[str, Any]] = []
        quarantine: list[dict[str, Any]] = []
        excluded: list[dict[str, Any]] = []
        issues: list[dict[str, Any]] = []
        candidates: list[dict[str, Any]] = []

        for sheet in sheets:
            profile = sheet.profile
            if profile.usage_class not in {CURRENT_PRODUCTION, HISTORICAL}:
                excluded.append(
                    {
                        "source_sheet": profile.source_sheet,
                        "usage_class": profile.usage_class,
                        "record_count": profile.row_count,
                        "admission_status": profile.admission_status,
                    }
                )
                continue
            for raw_record in sheet.records:
                payload = {policy.canonical_field(field): value for field, value in raw_record["values"].items()}
                source_columns = list(raw_record["values"])
                source_cells = {
                    policy.canonical_field(field): f"{_column_letter(index)}{int(raw_record['source_row'])}"
                    for index, field in enumerate(source_columns, start=1)
                }
                candidates.append(
                    self._quality_record(
                        payload,
                        source_file=profile.source_file,
                        source_file_hash=source_file_hash,
                        source_sheet=profile.source_sheet,
                        source_row=int(raw_record["source_row"]),
                        usage_class=profile.usage_class,
                        policy=policy,
                        imported_at=imported_at,
                        import_id=import_id,
                        previous_by_id=previous_by_id,
                        source_columns=source_columns,
                        source_cells=source_cells,
                    )
                )

        grouped: dict[str, list[dict[str, Any]]] = {}
        for record in candidates:
            grouped.setdefault(str(record["record_id"]), []).append(record)
        for record_id, group in grouped.items():
            fingerprints = {str(item["record_fingerprint"]) for item in group}
            if len(fingerprints) > 1:
                for record in group:
                    record["change_type"] = CONFLICT
                    record["admission_status"] = QUARANTINED
                    record["is_current"] = False
                    record["quality_issues"].append(
                        {"code": "duplicate_business_key_conflict", "severity": "critical", "record_id": record_id}
                    )
                    quarantine.append(record)
                issues.append({"code": "duplicate_business_key_conflict", "severity": "critical", "record_id": record_id})
                continue
            record = group[0]
            if len(group) > 1:
                record["quality_issues"].append({"code": "duplicate_identical_record", "severity": "medium"})
                record["quality_score"] = max(0, int(record["quality_score"]) - 5)
            if record["admission_status"] == QUARANTINED:
                quarantine.append(record)
                issues.extend(record["quality_issues"])
            elif record["admission_status"] == ADMITTED_CURRENT:
                current.append(record)
            else:
                historical.append(record)
        return current, historical, quarantine, excluded, issues

    def _quality_record(
        self,
        payload: dict[str, Any],
        *,
        source_file: str,
        source_file_hash: str,
        source_sheet: str,
        source_row: int,
        usage_class: str,
        policy: DataQualityPolicy,
        imported_at: str,
        import_id: str,
        previous_by_id: dict[str, dict[str, Any]],
        source_columns: list[str],
        source_cells: dict[str, str],
    ) -> dict[str, Any]:
        source_payload = dict(payload)
        payload, issues = self._normalize_payload(payload, policy)
        for required in policy.required_fields:
            if payload.get(required) in {None, ""}:
                issues.append({"code": "missing_required_field", "field": required, "severity": "high"})
        key_values = [payload.get(field_name) for field_name in policy.business_key_fields]
        if any(value in {None, ""} for value in key_values):
            issues.append({"code": "missing_business_key", "severity": "critical"})
            record_id = f"dq-invalid-{_sha256_value([source_file_hash, source_sheet, source_row])[:16]}"
        else:
            record_id = f"{policy.data_domain.lower()}-{_sha256_value([policy.data_domain, key_values])[:16]}"
        fingerprint = _sha256_value(payload)
        record_version_id = f"{record_id}-ver-{fingerprint[:12]}"
        previous = previous_by_id.get(record_id)
        if previous:
            change_type = UNCHANGED if str(previous.get("record_fingerprint")) == fingerprint else CHANGED
            created_time = str(previous.get("created_time") or imported_at)
        else:
            change_type = NEW
            created_time = imported_at
        status = str(payload.get(policy.status_field) or "")
        is_terminal = _normalized_text(status) in {_normalized_text(item) for item in policy.terminal_statuses}
        is_current = usage_class == CURRENT_PRODUCTION and not is_terminal and not issues
        admission_status = ADMITTED_CURRENT if is_current else ADMITTED_HISTORICAL
        if issues:
            admission_status = QUARANTINED
            is_current = False
        effective_value = payload.get(policy.effective_time_field)
        expire_value = payload.get(policy.expire_time_field)
        effective_time = _json_value(effective_value) if effective_value not in {None, ""} else imported_at
        expire_time = _json_value(expire_value) if expire_value not in {None, ""} else None
        if is_terminal and expire_time is None:
            expire_time = effective_time
        return {
            "schema_version": DATA_QUALITY_SCHEMA_VERSION,
            "record_id": record_id,
            "record_version_id": record_version_id,
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_row": source_row,
            "source_columns": source_columns,
            "source_cells": source_cells,
            "source_version": policy.source_version,
            "source_file_hash": source_file_hash,
            "record_fingerprint": fingerprint,
            "created_time": created_time,
            "effective_time": effective_time,
            "expire_time": expire_time,
            "status": status,
            "is_current": is_current,
            "change_type": change_type,
            "quality_score": max(0, 100 - len(issues) * 20),
            "quality_issues": issues,
            "admission_status": admission_status,
            "supersedes_record_id": record_id if previous and change_type == CHANGED else None,
            "supersedes_version_id": previous.get("record_version_id") if previous and change_type == CHANGED else None,
            "import_id": import_id,
            "adapter_id": policy.adapter_id,
            "mapping_version": policy.mapping_version,
            "data_domain": policy.data_domain,
            "source_payload": source_payload,
            "payload": payload,
        }

    @staticmethod
    def _normalize_payload(
        payload: dict[str, Any], policy: DataQualityPolicy
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        normalized: dict[str, Any] = {}
        issues: list[dict[str, Any]] = []
        for field_name, value in payload.items():
            field_type = policy.field_types.get(field_name, "")
            if value in {None, ""}:
                normalized[field_name] = value
                continue
            try:
                if field_type == "decimal":
                    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
                    normalized[field_name] = format(Decimal(text).quantize(Decimal("0.01")), "f")
                elif field_type == "date":
                    parsed = _parse_date_value(value)
                    if parsed is None:
                        raise ValueError("invalid date")
                    normalized[field_name] = parsed.date().isoformat()
                elif field_type == "string":
                    normalized[field_name] = str(value).strip()
                else:
                    normalized[field_name] = _json_value(value)
            except (InvalidOperation, ValueError):
                normalized[field_name] = _json_value(value)
                issues.append(
                    {
                        "code": "invalid_field_type",
                        "field": field_name,
                        "expected_type": field_type,
                        "severity": "high",
                    }
                )
        return normalized, issues

    @staticmethod
    def _missing_records(
        *,
        previous_records: list[dict[str, Any]],
        seen_ids: set[str],
        import_id: str,
        imported_at: str,
    ) -> list[dict[str, Any]]:
        missing: list[dict[str, Any]] = []
        for previous in previous_records:
            record_id = str(previous.get("record_id") or "")
            if not record_id or record_id in seen_ids or not previous.get("is_current"):
                continue
            record = dict(previous)
            record.update(
                {
                    "change_type": MISSING,
                    "admission_status": REVIEW_REQUIRED,
                    "is_current": False,
                    "import_id": import_id,
                    "observed_missing_at": imported_at,
                    "quality_issues": [
                        *(previous.get("quality_issues") or []),
                        {"code": "record_missing_from_latest_source", "severity": "high"},
                    ],
                }
            )
            missing.append(record)
        return missing

    @staticmethod
    def _counts(
        current: list[dict[str, Any]],
        historical: list[dict[str, Any]],
        quarantine: list[dict[str, Any]],
        missing: list[dict[str, Any]],
        excluded: list[dict[str, Any]],
    ) -> dict[str, int]:
        return {
            "current": len(current),
            "historical": len(historical),
            "quarantine": len(quarantine),
            "missing": len(missing),
            "excluded_sheets": len(excluded),
            "admitted": len(current) + len(historical),
        }

    def _audit(
        self,
        *,
        action: str,
        actor: Any,
        reason: str,
        result: str,
        target_id: str,
        correlation_id: str,
        metadata: dict[str, Any],
    ) -> dict[str, Any]:
        return self.audit.record(
            emp_id=actor.emp,
            actor_name=actor.name,
            module="data_quality",
            action=action,
            action_type=action,
            target_type="data_quality_import",
            target_id=target_id,
            reason=reason,
            result=result,
            correlation_id=correlation_id,
            metadata=metadata,
        )

    def _event(
        self,
        event_type: str,
        *,
        actor: Any,
        correlation_id: str,
        subject: str,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        dispatch = self.event_bus.publish(
            OMSEvent(
                event_type=event_type,
                source_module="data_quality",
                subject=subject,
                action=event_type,
                emp_id=actor.emp,
                actor_name=actor.name,
                correlation_id=correlation_id,
                payload=payload,
            )
        )
        return dict(dispatch["event"])

    def _attach_report(self, result: dict[str, Any]) -> dict[str, Any]:
        report_id = new_id("dqreport")
        safe_stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", Path(str(result.get("source_file") or "Excel")).stem)
        report_path = self.report_root / f"{safe_stem}_{result['import_id']}_数据质量报告.md"
        result["report_id"] = report_id
        result["report_path"] = str(report_path)
        self.report_writer.write(result, report_path)
        return result


class DataQualityReportWriter:
    """Write a deterministic, human-readable report for one import result."""

    def write(self, result: dict[str, Any], output_path: str | Path) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            f"# {result.get('source_file', 'Excel')}数据质量报告",
            "",
            f"- Import ID: `{result.get('import_id', '')}`",
            f"- Source Version: `{result.get('source_version', '')}`",
            f"- Source SHA-256: `{result.get('source_file_hash', '')}`",
            f"- Imported At: `{result.get('imported_at', '')}`",
            f"- Quality Status: `{result.get('quality_status', '')}`",
            "",
            "## Sheet 分析",
            "",
            "| Sheet | 行数 | 列数 | 时间范围 | 用途 | 准入状态 |",
            "|---|---:|---:|---|---|---|",
        ]
        for profile in result.get("sheet_profiles") or []:
            time_range = f"{profile.get('time_range_start') or '-'} -> {profile.get('time_range_end') or '-'}"
            lines.append(
                f"| {profile.get('source_sheet')} | {profile.get('row_count')} | {profile.get('column_count')} | "
                f"{time_range} | {profile.get('usage_class')} | {profile.get('admission_status')} |"
            )
        counts = result.get("counts") or {}
        lines.extend(
            [
                "",
                "## 准入结果",
                "",
                f"- 当前记录：{counts.get('current', 0)}",
                f"- 历史记录：{counts.get('historical', 0)}",
                f"- 隔离记录：{counts.get('quarantine', 0)}",
                f"- 缺失记录：{counts.get('missing', 0)}",
                f"- 排除 Sheet：{counts.get('excluded_sheets', 0)}",
                "",
                "## 异常记录",
                "",
            ]
        )
        issues = result.get("issues") or []
        if not issues:
            lines.append("- 无。")
        else:
            for issue in issues:
                lines.append(f"- `{issue.get('severity', 'unknown')}` {issue.get('code', 'unknown')}: {issue.get('message', '')}")
        lines.extend(
            [
                "",
                "## 结论",
                "",
                f"可进入 Truth Source 候选数量：{counts.get('admitted', 0)}。",
                "",
                "本报告只给出准入候选，不直接修改生产 Truth Source。",
            ]
        )
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return path


class TruthSourceSnapshotManager:
    """Create immutable acceptance snapshots and an optional active pointer."""

    def __init__(
        self,
        root: str | Path,
        *,
        audit: AuditEngine | None = None,
        event_bus: EventBus | None = None,
        master_data: OMSMasterData | None = None,
    ):
        self.root = Path(root)
        self.audit = audit or AuditEngine()
        self.event_bus = event_bus or EventBus()
        self.master_data = master_data or OMSMasterData()

    def create(
        self,
        *,
        acceptance_date: date | str,
        actor_emp_id: str,
        acceptance_run_id: str,
        acceptance_result: str,
        source_files: list[dict[str, Any]],
        source_sheets: list[dict[str, Any]],
        import_ids: list[str],
        imported_at: list[str],
        quality_report_ids: list[str],
        quality_results: dict[str, Any],
        adapter_versions: list[dict[str, Any]],
        truth_source_record_counts: dict[str, Any],
        metric_values: dict[str, Any],
        data_health_scores: dict[str, Any],
        snapshot_metadata: dict[str, Any] | None = None,
        hard_fail_reasons: list[str] | None = None,
        activate_for_production: bool = False,
        correlation_id: str = "",
    ) -> dict[str, Any]:
        if acceptance_result not in {HEALTH_PASS, HEALTH_WARNING, HEALTH_FAIL}:
            raise ValueError("acceptance_result must be PASS, WARNING, or FAIL")
        actor = self.master_data.employee_by_emp(actor_emp_id)
        if actor not in self.master_data.active_employees():
            raise PermissionError(f"inactive EMP cannot create a snapshot: {actor_emp_id}")
        day = acceptance_date.isoformat() if isinstance(acceptance_date, date) else str(acceptance_date)
        compact_day = day.replace("-", "")
        if not re.fullmatch(r"\d{8}", compact_day):
            raise ValueError("acceptance_date must be YYYY-MM-DD or YYYYMMDD")
        self.root.mkdir(parents=True, exist_ok=True)
        version = self._next_version(compact_day)
        previous = self._latest_version()
        activated = bool(activate_for_production and acceptance_result == HEALTH_PASS)
        if activate_for_production and acceptance_result != HEALTH_PASS:
            raise ValueError("only PASS snapshots can be activated for production")
        payload = {
            "schema_version": DATA_QUALITY_SCHEMA_VERSION,
            "snapshot_version": version,
            "acceptance_run_id": acceptance_run_id,
            "created_at": now_iso(),
            "created_by_emp_id": actor.emp,
            "source_files": source_files,
            "source_sheets": source_sheets,
            "import_ids": import_ids,
            "imported_at": imported_at,
            "quality_report_ids": quality_report_ids,
            "quality_results": quality_results,
            "adapter_versions": adapter_versions,
            "truth_source_record_counts": truth_source_record_counts,
            "metric_values": metric_values,
            "data_health_scores": data_health_scores,
            "snapshot_metadata": dict(snapshot_metadata or {}),
            "acceptance_result": acceptance_result,
            "hard_fail_reasons": list(hard_fail_reasons or []),
            "previous_snapshot_version": previous,
            "activated_for_production": activated,
            "immutable": True,
            "correlation_id": correlation_id or acceptance_run_id,
        }
        payload["snapshot_hash"] = _sha256_value(payload)
        target = self.root / f"{version}.json"
        if target.exists():
            raise FileExistsError(f"snapshot already exists: {version}")
        temporary = target.with_suffix(".json.tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        temporary.replace(target)
        if activated:
            pointer = {
                "snapshot_version": version,
                "snapshot_hash": payload["snapshot_hash"],
                "activated_at": now_iso(),
                "activated_by_emp_id": actor.emp,
            }
            pointer_path = self.root / "ACTIVE_SNAPSHOT.json"
            pointer_temp = pointer_path.with_suffix(".json.tmp")
            pointer_temp.write_text(json.dumps(pointer, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
            pointer_temp.replace(pointer_path)
        self.audit.record(
            emp_id=actor.emp,
            actor_name=actor.name,
            module="data_quality",
            action="data_quality.snapshot.created",
            action_type="create",
            target_type="truth_source_snapshot",
            target_id=version,
            reason="Create immutable Truth Source acceptance snapshot.",
            result=acceptance_result,
            correlation_id=correlation_id or acceptance_run_id,
            metadata={
                "snapshot_hash": payload["snapshot_hash"],
                "activated_for_production": activated,
            },
        )
        self.event_bus.publish(
            OMSEvent(
                event_type="data_quality.snapshot.available",
                source_module="data_quality",
                subject=version,
                action="snapshot_created",
                emp_id=actor.emp,
                actor_name=actor.name,
                correlation_id=correlation_id or acceptance_run_id,
                payload={
                    "snapshot_version": version,
                    "acceptance_result": acceptance_result,
                    "activated_for_production": activated,
                    "snapshot_hash": payload["snapshot_hash"],
                },
            )
        )
        return payload

    def read(self, snapshot_version: str) -> dict[str, Any]:
        path = self.root / f"{snapshot_version}.json"
        if not path.exists():
            raise KeyError(f"unknown Truth Source snapshot: {snapshot_version}")
        payload = json.loads(path.read_text(encoding="utf-8"))
        expected_hash = str(payload.pop("snapshot_hash", ""))
        actual_hash = _sha256_value(payload)
        payload["snapshot_hash"] = expected_hash
        if expected_hash != actual_hash:
            raise ValueError(f"snapshot integrity check failed: {snapshot_version}")
        return payload

    def active(self) -> dict[str, Any] | None:
        path = self.root / "ACTIVE_SNAPSHOT.json"
        if not path.exists():
            return None
        pointer = json.loads(path.read_text(encoding="utf-8"))
        return self.read(str(pointer["snapshot_version"]))

    def _next_version(self, compact_day: str) -> str:
        pattern = re.compile(rf"^TS-{compact_day}-V(\d+)\.json$")
        numbers = []
        for path in self.root.glob(f"TS-{compact_day}-V*.json"):
            match = pattern.match(path.name)
            if match:
                numbers.append(int(match.group(1)))
        return f"TS-{compact_day}-V{max(numbers, default=0) + 1}"

    def _latest_version(self) -> str | None:
        versions = []
        pattern = re.compile(r"^TS-(\d{8})-V(\d+)\.json$")
        for path in self.root.glob("TS-*-V*.json"):
            match = pattern.match(path.name)
            if match:
                versions.append((match.group(1), int(match.group(2)), path.stem))
        return max(versions)[2] if versions else None


def default_quality_policy(data_domain: str, *, source_version: str) -> DataQualityPolicy:
    domain = data_domain.strip().lower()
    policies: dict[str, dict[str, Any]] = {
        "sales": {
            "required_fields": ("customer_name", "amount"),
            "business_key_fields": ("contract_id",),
            "status_field": "status",
            "effective_time_field": "contract_date",
            "field_aliases": {
                "customer_name": ("客户姓名", "宝妈姓名", "姓名", "客户"),
                "contract_id": ("合同号", "合同编号", "contract_no"),
                "amount": ("合同金额", "成交金额", "全款费用", "金额"),
                "status": ("销售状态", "合同状态", "阶段"),
                "contract_date": ("签约日期", "合同日期"),
                "salesperson": ("销售", "销售人员", "销售顾问"),
                "expected_delivery_date": ("预产期", "预计生产日期"),
                "actual_received_amount": ("实际到账金额", "实收金额"),
            },
            "field_types": {
                "customer_name": "string",
                "contract_id": "string",
                "amount": "decimal",
                "status": "string",
                "contract_date": "date",
                "actual_received_amount": "decimal",
            },
        },
        "finance": {
            "required_fields": ("tx_id", "amount", "type"),
            "business_key_fields": ("tx_id",),
            "status_field": "status",
            "effective_time_field": "tx_date",
            "field_aliases": {
                "tx_id": ("流水号", "交易编号", "financial_event_id"),
                "amount": ("金额", "收入金额", "支出金额"),
                "type": ("类型", "收支类型", "方向"),
                "status": ("支付状态", "对账状态", "状态"),
                "tx_date": ("日期", "发生日期", "交易日期"),
            },
            "field_types": {
                "tx_id": "string",
                "amount": "decimal",
                "type": "string",
                "status": "string",
                "tx_date": "date",
            },
        },
        "room": {
            "required_fields": ("room_id", "status"),
            "business_key_fields": ("room_id",),
            "status_field": "status",
            "field_aliases": {
                "room_id": ("房间号", "房号"),
                "status": ("房态", "房间状态", "状态"),
                "customer_name": ("客户姓名", "姓名", "当前客户"),
                "stay_id": ("入住编号", "stay_id"),
            },
            "field_types": {"room_id": "string", "status": "string", "stay_id": "string"},
        },
        "stay": {
            "required_fields": ("stay_id", "customer_name", "checkin_date"),
            "business_key_fields": ("stay_id",),
            "status_field": "status",
            "effective_time_field": "checkin_date",
            "expire_time_field": "checkout_date",
            "field_aliases": {
                "stay_id": ("入住编号", "记录编号"),
                "customer_name": ("客户姓名", "姓名"),
                "room_id": ("房间号", "房号"),
                "checkin_date": ("入住日期", "入住时间"),
                "checkout_date": ("出馆日期", "预计出馆日期"),
                "status": ("入住状态", "状态"),
            },
            "field_types": {
                "stay_id": "string",
                "customer_name": "string",
                "room_id": "string",
                "checkin_date": "date",
                "checkout_date": "date",
                "status": "string",
            },
        },
        "contract": {
            "required_fields": ("contract_id", "customer_name", "amount"),
            "business_key_fields": ("contract_id",),
            "status_field": "status",
            "effective_time_field": "contract_date",
            "field_aliases": {
                "contract_id": ("合同号", "合同编号"),
                "customer_name": ("客户姓名", "姓名"),
                "amount": ("合同金额", "金额"),
                "contract_date": ("签约日期", "合同日期"),
                "status": ("合同状态", "状态"),
            },
            "field_types": {
                "contract_id": "string",
                "customer_name": "string",
                "amount": "decimal",
                "contract_date": "date",
                "status": "string",
            },
        },
        "customer": {
            "required_fields": ("customer_id", "customer_name"),
            "business_key_fields": ("customer_id",),
            "status_field": "status",
            "field_aliases": {
                "customer_id": ("客户编号", "客户id"),
                "customer_name": ("客户姓名", "姓名"),
                "status": ("客户状态", "状态"),
                "expected_checkin_date": ("预计入住日期", "预产期"),
            },
            "field_types": {
                "customer_id": "string",
                "customer_name": "string",
                "status": "string",
                "expected_checkin_date": "date",
            },
        },
    }
    try:
        config = policies[domain]
    except KeyError as exc:
        raise KeyError(f"no default data quality policy for domain: {data_domain}") from exc
    return DataQualityPolicy(
        data_domain=data_domain.title(),
        source_version=source_version,
        adapter_id=f"{domain}_adapter_v1",
        mapping_version="p0.14.data_quality.v1",
        required_fields=config["required_fields"],
        business_key_fields=config["business_key_fields"],
        field_aliases=config["field_aliases"],
        field_types=config.get("field_types", {}),
        status_field=config.get("status_field", "status"),
        effective_time_field=config.get("effective_time_field", "effective_time"),
        expire_time_field=config.get("expire_time_field", "expire_time"),
    )
