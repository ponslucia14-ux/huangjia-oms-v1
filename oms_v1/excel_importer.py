from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any

from .live_connector import DEFAULT_LIVE_ROOT
from .operating_center_source import OPERATING_CENTER_PEOPLE, OPERATING_CENTER_VERSION
from .schemas import new_id, now_iso


EXCEL_IMPORT_SCHEMA_VERSION = "oms.v1.excel_import_stream"
EXCEL_LEGACY_POLICY = {
    "Excel": "input_source",
    "微信": "input_source_only",
    "OMS": "default_work_entry",
}

EXCEL_METADATA_KEYS = {"__source_sheet", "__row_number"}
XLSX_SCAN_MAX_COLUMNS = 256
XLSX_HEADER_SCAN_ROWS = 30
RECOGNIZED_HEADERS = {
    "序号",
    "签约日期",
    "客户",
    "客户姓名",
    "姓名",
    "妈妈姓名",
    "签约客户",
    "电话",
    "套餐",
    "顾问",
    "价格",
    "金额",
    "合同",
    "合同号",
    "合同编号",
    "订单号",
    "房间",
    "房间号",
    "房号",
    "房型",
    "房态",
    "预产期",
    "生产时间",
    "入住时间",
    "入住日期",
    "出馆时间",
    "天数",
    "销售",
    "管家",
    "照护师",
    "服务",
    "备注",
    "需求",
}


SOURCE_CONFIG = {
    "resident": {
        "source_name": "在住表",
        "workspace_key": "nana",
        "layer": "business_layer",
        "system_capabilities": ["风险预警", "服务安排"],
        "daily_process": "入住服务跟进",
        "action_type": "excel_resident_service_task",
        "sync_target": "Excel_在住表",
    },
    "room_status": {
        "source_name": "房态表",
        "workspace_key": "june",
        "layer": "business_layer",
        "system_capabilities": ["排房优化", "房态冲突识别"],
        "daily_process": "房态排房处理",
        "action_type": "excel_room_status_task",
        "sync_target": "Excel_房态表",
    },
    "contracts": {
        "source_name": "签约客户表",
        "workspace_key": "huanhuan",
        "layer": "business_layer",
        "system_capabilities": ["客户结构化", "经营指标中心"],
        "daily_process": "签约客户提报",
        "action_type": "excel_contract_customer_task",
        "sync_target": "Excel_签约客户表",
    },
}

MEANINGFUL_VALUE_EXCLUDE_KEYS = {"序号"}


class ExcelOMSImporter:
    """Convert real Excel/CSV rows into OMS user-centric work items."""

    def __init__(self, live_root: str | Path | None = None, operating_root: str | Path | None = None):
        self.live_root = Path(live_root or os.getenv("OMS_LIVE_ROOT") or DEFAULT_LIVE_ROOT)
        self.operating_root = Path(operating_root or self.live_root / "operational_core")

    def import_sources(
        self,
        *,
        resident: str | Path | None = None,
        room_status: str | Path | None = None,
        contracts: str | Path | None = None,
    ) -> dict[str, Any]:
        source_paths = {
            "resident": resident,
            "room_status": room_status,
            "contracts": contracts,
        }
        records: list[dict[str, Any]] = []
        errors: list[dict[str, str]] = []
        for source_type, path in source_paths.items():
            if not path:
                continue
            try:
                rows = self._read_table(Path(path))
            except Exception as exc:
                errors.append({"source_type": source_type, "path": str(path), "error": str(exc)})
                continue
            records.extend(self._record(source_type, row, index + 1, Path(path)) for index, row in enumerate(rows))

        work_items = [self._work_item(record) for record in records]
        pending = [self._pending_outbox(record, work_item) for record, work_item in zip(records, work_items)]
        self._persist_work_items(work_items)
        self._persist_pending(pending)
        return {
            "schema_version": EXCEL_IMPORT_SCHEMA_VERSION,
            "source_of_truth": "Excel",
            "people_model_source": OPERATING_CENTER_VERSION,
            "input_sources": {key: str(path) for key, path in source_paths.items() if path},
            "record_count": len(records),
            "work_item_count": len(work_items),
            "pending_outbox_count": len(pending),
            "records": records,
            "work_items": work_items,
            "pending_outbox": pending,
            "errors": errors,
            "audit": {
                "created_at": now_iso(),
                "live_root": str(self.live_root),
                "operating_root": str(self.operating_root),
                "principle": "Excel数据进入 user_id -> workspace -> role 的个人工作台模型。",
            },
        }

    def _read_table(self, path: Path) -> list[dict[str, Any]]:
        if not path.exists():
            raise FileNotFoundError(f"Excel source not found: {path}")
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                return [
                    {**self._clean_row(row), "__row_number": index + 2, "__source_sheet": ""}
                    for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter))
                ]
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx(path)
        raise ValueError(f"Unsupported Excel source type: {suffix}")

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read .xlsx sources in this runtime") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                result.extend(self._read_xlsx_sheet(sheet))
            return result
        finally:
            workbook.close()

    def _read_xlsx_sheet(self, sheet: Any) -> list[dict[str, Any]]:
        max_col = min(sheet.max_column or 1, XLSX_SCAN_MAX_COLUMNS)
        scan_rows = list(
            sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row or 1, XLSX_HEADER_SCAN_ROWS),
                max_col=max_col,
                values_only=True,
            )
        )
        header_position = self._detect_header_row(scan_rows)
        if header_position is None:
            return []

        header_values = scan_rows[header_position - 1]
        headers = self._headers(header_values)
        result: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_position + 1, max_col=max_col, values_only=True),
            start=header_position + 1,
        ):
            row = {
                headers[index]: values[index] if index < len(values) else ""
                for index in range(len(headers))
                if headers[index]
            }
            cleaned = self._clean_row(row)
            if not self._is_business_data_row(cleaned):
                continue
            cleaned["__source_sheet"] = sheet.title
            cleaned["__row_number"] = row_number
            result.append(cleaned)
        return result

    def _detect_header_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        best: tuple[int, int] | None = None
        for index, values in enumerate(rows, start=1):
            labels = [str(value).strip() for value in values if value not in {"", None}]
            hits = sum(1 for label in labels if label in RECOGNIZED_HEADERS)
            if hits < 2:
                continue
            score = hits * 10 + len(labels)
            if best is None or score > best[1]:
                best = (index, score)
        return best[0] if best else None

    def _headers(self, values: tuple[Any, ...]) -> list[str]:
        headers: list[str] = []
        seen: dict[str, int] = {}
        for index, value in enumerate(values, start=1):
            header = str(value or "").strip()
            if not header:
                headers.append("")
                continue
            count = seen.get(header, 0)
            seen[header] = count + 1
            headers.append(header if count == 0 else f"{header}_{count + 1}")
        return headers

    def _is_business_data_row(self, row: dict[str, Any]) -> bool:
        values = [value for key, value in row.items() if key not in EXCEL_METADATA_KEYS and value not in {"", None}]
        if not values:
            return False
        meaningful_values = [
            value
            for key, value in row.items()
            if key not in EXCEL_METADATA_KEYS
            and key not in MEANINGFUL_VALUE_EXCLUDE_KEYS
            and value not in {"", None}
            and str(value).strip()
        ]
        if not meaningful_values:
            return False
        header_like_hits = sum(1 for value in values if str(value).strip() in RECOGNIZED_HEADERS)
        return header_like_hits < max(2, len(values) // 2)

    def _clean_row(self, row: dict[str, Any]) -> dict[str, Any]:
        return {str(key).strip(): "" if value is None else value for key, value in row.items() if str(key).strip()}

    def _record(self, source_type: str, row: dict[str, Any], row_number: int, path: Path) -> dict[str, Any]:
        config = SOURCE_CONFIG[source_type]
        person = self._person(config["workspace_key"])
        user_id = os.getenv(person["feishu_env"], "").strip()
        status = "mapped" if user_id else "unresolved_user_id"
        source_sheet = str(row.get("__source_sheet", ""))
        source_row_number = int(row.get("__row_number") or row_number)
        business_row = {key: value for key, value in row.items() if key not in EXCEL_METADATA_KEYS}
        return {
            "record_id": new_id("excel"),
            "source_type": source_type,
            "source_name": config["source_name"],
            "source_file": str(path),
            "source_sheet": source_sheet,
            "row_number": source_row_number,
            "raw_row": business_row,
            "normalized": self._normalized_fields(source_type, business_row),
            "assignment": {
                "user_id": user_id,
                "user_id_status": status,
                "workspace_key": config["workspace_key"],
                "workspace": person["title"],
                "role": person["role"],
                "name": person["name"],
            },
            "structure": {
                "layer": config["layer"],
                "support_layer": self._support_layer(source_type, row),
                "system_capability_layer": config["system_capabilities"],
            },
        }

    def _work_item(self, record: dict[str, Any]) -> dict[str, Any]:
        config = SOURCE_CONFIG[record["source_type"]]
        assignment = record["assignment"]
        status = "ready_with_pending_sync" if assignment["user_id"] else "attention_required"
        return {
            "schema_version": "oms.v1.operational_work_item",
            "work_item_id": new_id("op"),
            "role": assignment["role"],
            "workspace": assignment["workspace"],
            "daily_process": config["daily_process"],
            "primary_entry": "OMS",
            "legacy_policy": EXCEL_LEGACY_POLICY,
            "action_id": record["record_id"],
            "action_type": config["action_type"],
            "status": status,
            "confirmation_required": not bool(assignment["user_id"]),
            "next_operator_action": self._next_action(record),
            "source_sync_targets": [config["sync_target"], "OMS pending_outbox"],
            "excel_record": record,
        }

    def _pending_outbox(self, record: dict[str, Any], work_item: dict[str, Any]) -> dict[str, Any]:
        return {
            "created_at": now_iso(),
            "mode": "Excel_Pending_Mode",
            "source_of_truth": "Excel",
            "people_model_source": OPERATING_CENTER_VERSION,
            "status": "pending",
            "reason": "Excel row converted to OMS work_item; external sync waits for live connector availability.",
            "record": record,
            "work_item": work_item,
        }

    def _normalized_fields(self, source_type: str, row: dict[str, Any]) -> dict[str, Any]:
        field_synonyms = {
            "customer_name": ["客户", "客户姓名", "姓名", "妈妈姓名", "签约客户"],
            "room": ["房间", "房间号", "房号", "房型", "房态"],
            "checkin_date": ["入住日期", "入住", "预产期", "到店日期"],
            "contract_no": ["合同", "合同号", "合同编号", "订单号"],
            "amount": ["金额", "价格", "收款", "定金", "尾款", "合同金额"],
            "service_note": ["服务", "备注", "需求", "特殊餐", "护理"],
        }
        normalized = {field: self._first(row, names) for field, names in field_synonyms.items()}
        normalized["business_type"] = source_type
        return normalized

    def _first(self, row: dict[str, Any], names: list[str]) -> str:
        for name in names:
            if name in row and row[name] not in {"", None}:
                return str(row[name])
        return ""

    def _support_layer(self, source_type: str, row: dict[str, Any]) -> list[str]:
        joined = " ".join(str(value) for value in row.values())
        support: list[str] = []
        if any(word in joined for word in ["厨房", "餐", "食材", "特殊餐"]):
            support.append("餐饮/厨房")
        if any(word in joined for word in ["产护", "护理", "月嫂"]):
            support.append("产护支持")
        if any(word in joined for word in ["清理", "维修", "设备", "物资"]):
            support.append("后勤保障")
        if any(word in joined for word in ["采购", "消耗品", "报销"]):
            support.append("行政采购")
        if source_type == "resident" and not support:
            support.extend(["产护支持", "餐饮/厨房"])
        return support

    def _next_action(self, record: dict[str, Any]) -> str:
        assignment = record["assignment"]
        if assignment["user_id"]:
            return f"进入 {assignment['workspace']}，由 {assignment['name']} 处理 Excel 来源事项。"
        return f"{assignment['workspace']} 缺少真实飞书 user_id，事项保留在 pending_outbox。"

    def _person(self, workspace_key: str) -> dict[str, Any]:
        return OPERATING_CENTER_PEOPLE[workspace_key]

    def _persist_work_items(self, work_items: list[dict[str, Any]]) -> None:
        self.operating_root.mkdir(parents=True, exist_ok=True)
        path = self.operating_root / "excel_work_items.jsonl"
        with path.open("a", encoding="utf-8") as handle:
            for item in work_items:
                handle.write(json.dumps(item, ensure_ascii=False) + "\n")

    def _persist_pending(self, pending: list[dict[str, Any]]) -> None:
        path = self.live_root / "pending_outbox" / "Excel_OMS导入.jsonl"
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            for item in pending:
                handle.write(json.dumps(item, ensure_ascii=False) + "\n")
