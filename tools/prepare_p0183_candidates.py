from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from oms_v1.data_quality import HEALTH_PASS, HEALTH_WARNING, SheetSemanticMemory


CUTOVER_DATE = date(2026, 7, 11)


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def number(value: Any) -> float | None:
    if value in (None, "", "无"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace("¥", "").replace("￥", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date(value: Any, *, default_year: int = 2026) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 30000:
        try:
            return from_excel(value).date()
        except (TypeError, ValueError, OverflowError):
            return None
    raw = str(value or "").strip().replace("年", ".").replace("月", ".").replace("日", "")
    raw = raw.replace("/", ".").replace("-", ".")
    parts = [item for item in raw.split(".") if item]
    try:
        if len(parts) >= 3:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        if len(parts) == 2:
            return date(default_year, int(parts[0]), int(parts[1]))
    except ValueError:
        return None
    return None


def workbook_inventory(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        result: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            nonempty_rows = 0
            for row in sheet.iter_rows(values_only=True):
                if any(value not in (None, "") for value in row):
                    nonempty_rows += 1
            result.append(
                {
                    "source_sheet": sheet.title,
                    "row_count": nonempty_rows,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "visibility": str(sheet.sheet_state or "visible").upper(),
                }
            )
        return result
    finally:
        workbook.close()


def sales_candidates(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    target_sheets = ["2026年内店销售列表", "2026外店上户列表"]
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    per_sheet: dict[str, int] = {}
    try:
        for sheet_name in target_sheets:
            sheet = workbook[sheet_name]
            headers: dict[str, int] = {}
            count = 0
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(row)
                if "合同编号" in values:
                    headers = {str(value).strip(): index for index, value in enumerate(values) if value not in (None, "")}
                    continue
                if not headers:
                    continue
                contract_index = headers.get("合同编号")
                contract_id = text(values[contract_index]) if contract_index is not None and contract_index < len(values) else ""
                if not contract_id.startswith("NSEKI"):
                    continue
                def field(name: str) -> Any:
                    index = headers.get(name)
                    return values[index] if index is not None and index < len(values) else None

                signed_at = parse_date(field("签约日期"))
                customer_name = text(field("宝妈姓名"))
                amount = number(field("全款费用"))
                salesperson = text(field("销售"))
                missing = [
                    name for name, value in {
                        "contract_id": contract_id,
                        "signed_at": signed_at,
                        "customer_name": customer_name,
                        "amount": amount,
                    }.items() if value in (None, "")
                ]
                if missing:
                    issues.append({"sheet": sheet_name, "row": row_number, "issue": "required_fields_missing", "fields": missing})
                    continue
                records.append(
                    {
                        "record_id": f"sales:{contract_id}",
                        "source_file": path.name,
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                        "contract_id": contract_id,
                        "signed_at": signed_at.isoformat(),
                        "customer_name": customer_name,
                        "amount": amount,
                        "salesperson": salesperson,
                        "classification": "CURRENT_CANDIDATE",
                        "current_basis": "signed_contract_business_scope",
                    }
                )
                count += 1
            per_sheet[sheet_name] = count
    finally:
        workbook.close()

    counts: dict[str, int] = {}
    for item in records:
        counts[item["contract_id"]] = counts.get(item["contract_id"], 0) + 1
    duplicate_ids = sorted(contract_id for contract_id, count in counts.items() if count > 1)
    if duplicate_ids:
        issues.append({"issue": "duplicate_contract_id", "values": duplicate_ids})
    quarantined = [item for item in records if item["contract_id"] in duplicate_ids]
    records = [item for item in records if item["contract_id"] not in duplicate_ids]
    for item in quarantined:
        item["classification"] = "CONFLICT"
    return {
        "domain": "Sales",
        "source_file": path.name,
        "source_sha256": file_hash(path),
        "classification": "CONDITIONAL_CURRENT",
        "quality_result": HEALTH_WARNING if issues else HEALTH_PASS,
        "records": records,
        "record_count": len(records),
        "quarantine_records": quarantined,
        "quarantine_count": len(quarantined),
        "per_sheet": per_sheet,
        "issues": issues,
        "blocking_reason": "contract lifecycle status is not explicit" if not issues else "record quality issues exist",
    }


def finance_historical(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    per_sheet: dict[str, int] = {}
    try:
        for sheet in workbook.worksheets:
            if sheet.title == "日结":
                continue
            count = 0
            for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
                values = list(row)
                layouts = [(0, 1, 2, 3, 4, "凰家"), (8 if sheet.max_column >= 14 else 7, 9 if sheet.max_column >= 14 else 8, 10 if sheet.max_column >= 14 else 9, 11 if sheet.max_column >= 14 else 10, 12 if sheet.max_column >= 14 else 11, "凤稚")]
                for date_i, income_item_i, income_amount_i, expense_item_i, expense_amount_i, entity in layouts:
                    if date_i >= len(values):
                        continue
                    occurred_at = parse_date(values[date_i])
                    if occurred_at is None:
                        continue
                    for direction, item_i, amount_i in (("INCOME", income_item_i, income_amount_i), ("EXPENSE", expense_item_i, expense_amount_i)):
                        item = text(values[item_i]) if item_i < len(values) else ""
                        amount = number(values[amount_i]) if amount_i < len(values) else None
                        if not item and amount is None:
                            continue
                        if not item or amount is None:
                            issues.append({"sheet": sheet.title, "row": row_number, "entity": entity, "direction": direction, "issue": "item_amount_pair_incomplete"})
                            continue
                        key = f"{sheet.title}|{row_number}|{entity}|{direction}"
                        records.append(
                            {
                                "record_id": f"finance:{hashlib.sha256(key.encode()).hexdigest()[:20]}",
                                "source_file": path.name,
                                "source_sheet": sheet.title,
                                "source_row": row_number,
                                "business_entity": entity,
                                "occurred_at": occurred_at.isoformat(),
                                "direction": direction,
                                "item": item,
                                "amount": amount,
                                "classification": "HISTORICAL",
                            }
                        )
                        count += 1
            per_sheet[sheet.title] = count
    finally:
        workbook.close()
    return {
        "domain": "Finance",
        "source_file": path.name,
        "source_sha256": file_hash(path),
        "classification": "HISTORICAL_ONLY",
        "quality_result": HEALTH_WARNING if issues else HEALTH_PASS,
        "records": records,
        "record_count": len(records),
        "quarantine_count": len(issues),
        "per_sheet": per_sheet,
        "excluded_sheets": ["日结"],
        "issues": issues,
        "current_generation": "OMS operation by EMP004, reviewed by EMP003",
    }


def contract_plan(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    per_sheet: dict[str, int] = {}
    try:
        for sheet in workbook.worksheets:
            headers: dict[str, int] = {}
            count = 0
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(row)
                if "签约日期" in values and "姓名" in values:
                    headers = {str(value).strip(): index for index, value in enumerate(values) if value not in (None, "")}
                    continue
                if not headers:
                    continue
                def field(*names: str) -> Any:
                    for name in names:
                        index = headers.get(name)
                        if index is not None and index < len(values):
                            return values[index]
                    return None

                sequence = number(field("序号"))
                customer_name = text(field("姓名"))
                expected_date = parse_date(field("预产期"))
                signed_at = parse_date(field("签约日期"))
                package = text(field("套餐"))
                if sequence is None or not customer_name:
                    continue
                if expected_date is None or not package:
                    issues.append({"sheet": sheet.title, "row": row_number, "issue": "plan_fields_missing"})
                    continue
                record_key = f"{sheet.title}|{row_number}|{customer_name}|{expected_date.isoformat()}"
                records.append(
                    {
                        "record_id": f"contract-plan:{hashlib.sha256(record_key.encode()).hexdigest()[:20]}",
                        "source_file": path.name,
                        "source_sheet": sheet.title,
                        "source_row": row_number,
                        "customer_name": customer_name,
                        "signed_at": signed_at.isoformat() if signed_at else None,
                        "expected_date": expected_date.isoformat(),
                        "package": package,
                        "advisor": text(field("顾问", "销售")),
                        "classification": "CONTRACT_STAY_PLAN",
                        "plan_time_scope": "UPCOMING" if expected_date >= CUTOVER_DATE else "HISTORICAL_PLAN",
                    }
                )
                count += 1
            per_sheet[sheet.title] = count
    finally:
        workbook.close()
    return {
        "domain": "ContractStayPlan",
        "source_file": path.name,
        "source_sha256": file_hash(path),
        "classification": "PLAN",
        "quality_result": HEALTH_WARNING if issues else HEALTH_PASS,
        "records": records,
        "record_count": len(records),
        "quarantine_count": len(issues),
        "upcoming_count": sum(item["plan_time_scope"] == "UPCOMING" for item in records),
        "historical_plan_count": sum(item["plan_time_scope"] == "HISTORICAL_PLAN" for item in records),
        "per_sheet": per_sheet,
        "issues": issues,
        "actual_stay_prohibited": True,
    }


def actual_stay_candidate(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    sheet = workbook["房态表"]
    headers = {str(value).strip(): index for index, value in enumerate(next(sheet.iter_rows(values_only=True))) if value not in (None, "")}
    update_date: date | None = None
    reported_resident_count: int | None = None
    try:
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            values = list(row)
            for value in values:
                match = re.search(r"更新时间[:：]\s*(\d{4})\.(\d{1,2})\.(\d{1,2})", str(value or ""))
                if match:
                    update_date = date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
                resident_match = re.search(r"产妇数量[:：]\s*(\d+)人", str(value or ""))
                if resident_match:
                    reported_resident_count = int(resident_match.group(1))
            room_id = text(values[headers["房间号"]]) if headers.get("房间号") is not None and headers["房间号"] < len(values) else ""
            customer_name = text(values[headers["姓名"]]) if headers.get("姓名") is not None and headers["姓名"] < len(values) else ""
            if not room_id or not customer_name or not re.match(r"^(\d{3}|伟峰\d+)", room_id):
                continue
            checkin = parse_date(values[headers["入住时间"]]) if headers.get("入住时间") is not None else None
            checkout = parse_date(values[headers["出馆时间"]]) if headers.get("出馆时间") is not None else None
            if checkin is None or checkout is None:
                issues.append({"sheet": sheet.title, "row": row_number, "room_id": room_id, "customer_name": customer_name, "issue": "actual_dates_missing"})
                continue
            record_key = f"{room_id}|{customer_name}|{checkin.isoformat()}"
            records.append(
                {
                    "record_id": f"actual-stay:{hashlib.sha256(record_key.encode()).hexdigest()[:20]}",
                    "source_file": path.name,
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                    "room_id": room_id,
                    "customer_name": customer_name,
                    "checkin_date": checkin.isoformat(),
                    "checkout_date": checkout.isoformat(),
                    "butler": text(values[headers["管家"]]) if headers.get("管家") is not None else "",
                    "caregiver": text(values[headers["照护师"]]) if headers.get("照护师") is not None else "",
                    "classification": "ACTUAL_STAY_HISTORICAL_SNAPSHOT",
                }
            )
    finally:
        workbook.close()
    stale = update_date is None or update_date < CUTOVER_DATE
    active_at_source_count = sum(
        parse_date(item["checkin_date"]) <= update_date <= parse_date(item["checkout_date"])
        for item in records
        if update_date is not None
    )
    if reported_resident_count is not None and active_at_source_count != reported_resident_count:
        issues.append(
            {
                "issue": "resident_summary_mismatch",
                "effective_date": update_date.isoformat() if update_date else None,
                "calculated_count": active_at_source_count,
                "reported_count": reported_resident_count,
            }
        )
    return {
        "domain": "ActualStay",
        "source_file": path.name,
        "source_sha256": file_hash(path),
        "source_sheet": "房态表",
        "source_effective_date": update_date.isoformat() if update_date else None,
        "classification": "HISTORICAL_SNAPSHOT" if stale else "CURRENT_CANDIDATE",
        "quality_result": HEALTH_WARNING if stale or issues else HEALTH_PASS,
        "records": records,
        "record_count": len(records),
        "quarantine_count": sum(item.get("issue") == "actual_dates_missing" for item in issues),
        "active_at_source_count": active_at_source_count,
        "reported_resident_count": reported_resident_count,
        "issues": issues + ([{"issue": "source_effective_date_before_cutover", "effective_date": update_date.isoformat() if update_date else None}] if stale else []),
        "current_generation": "OMS operation by EMP008",
    }


def profile(sheet: dict[str, Any], fields: Iterable[str]) -> dict[str, Any]:
    return {
        "header_row": 1,
        "column_count": sheet["max_column"],
        "fields": [{"source_field": item, "canonical_field": item} for item in fields],
    }


def create_memories(output: Path, sources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    memory = SheetSemanticMemory(output / "sheet_semantic_memory_candidates.json")
    entries: list[dict[str, Any]] = []
    definitions = [
        (sources[0], "2026年销售明细表（经验为王*.xlsx", ["2026年内店销售列表", "2026外店上户列表"], "Sales", "Sales Current Candidate", "EMP004", ["合同编号", "签约日期", "宝妈姓名", "全款费用"]),
        (sources[1], "2026年财务报表（7月）*.xlsx", [item["source_sheet"] for item in sources[1]["inventory"] if item["source_sheet"] != "日结"], "Finance", "Finance Historical", "EMP004", ["日期", "收入项目", "收入金额", "支出项目", "支出金额"]),
        (sources[2], "A  凰家母婴签约客户一览表（挤牙膏）*.xlsx", [item["source_sheet"] for item in sources[2]["inventory"]], "ContractStayPlan", "Contract Stay Plan", "EMP008", ["签约日期", "姓名", "预产期", "套餐"]),
        (sources[3], "凰家母婴在住表*.xlsx", ["房态表"], "ActualStay", "Actual Stay Historical Snapshot", "EMP009", ["房间号", "姓名", "入住时间", "出馆时间"]),
    ]
    for source, pattern, sheets, domain, fact_type, owner, fields in definitions:
        inventory = {item["source_sheet"]: item for item in source["inventory"]}
        for sheet_name in sheets:
            sheet = inventory[sheet_name]
            entries.append(
                memory.confirm(
                    source_file_pattern=pattern,
                    source_sheet=sheet_name,
                    domain=domain,
                    fact_type=fact_type,
                    owner=owner,
                    profile=profile(sheet, fields),
                    workbook_sheets=[item["source_sheet"] for item in source["inventory"]],
                    quality_result=source["analysis"]["quality_result"],
                    memory_status=SheetSemanticMemory.TEMPORARY,
                    reason="P0.18.3 business evidence candidate; requires Data Quality acceptance before confirmation.",
                )
            )
    return entries


def find_sources(source_root: Path) -> tuple[Path, Path, Path, Path]:
    files = list(source_root.glob("*.xlsx"))
    sales = next(path for path in files if "7.10" in path.name and "(1)" in path.name and path.stat().st_size == 226352)
    finance = next(path for path in files if path.stat().st_size == 72501)
    contract = next(path for path in files if path.stat().st_size == 352020)
    actual_stay = next(path for path in files if path.stat().st_size == 50001)
    return sales, finance, contract, actual_stay


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    paths = find_sources(args.source_root)
    analyses = [sales_candidates(paths[0]), finance_historical(paths[1]), contract_plan(paths[2]), actual_stay_candidate(paths[3])]
    sources: list[dict[str, Any]] = []
    for path, analysis in zip(paths, analyses):
        item = {"source_file": path.name, "source_path": str(path), "inventory": workbook_inventory(path), "analysis": analysis}
        sources.append(item)
        (args.output / f"{analysis['domain']}_data_quality_candidate.json").write_text(
            json.dumps(item, ensure_ascii=False, indent=2, default=json_value), encoding="utf-8"
        )
    memories = create_memories(args.output, sources)
    preparation = {
        "schema_version": "oms.p0183.v2_preparation",
        "candidate_snapshot": "TS-20260711-V2",
        "status": "PREPARED_NOT_GENERATED",
        "active_snapshot_unchanged": "TS-20260711-V1",
        "generated_at": datetime.now().astimezone().isoformat(),
        "domains": {
            analysis["domain"]: {
                "source_file": analysis["source_file"],
                "source_sha256": analysis["source_sha256"],
                "classification": analysis["classification"],
                "quality_result": analysis["quality_result"],
                "record_count": analysis["record_count"],
                "quarantine_count": analysis.get("quarantine_count", 0),
                "issue_count": len(analysis["issues"]),
            }
            for analysis in analyses
        },
        "semantic_memory_candidate_count": len(memories),
        "semantic_memory_status": "TEMPORARY",
        "generation_gate": {
            "requires_all_current": False,
            "requires_quality_acceptance": True,
            "requires_no_false_current": True,
            "ready": False,
        },
    }
    (args.output / "TS-20260711-V2_PREPARATION.json").write_text(
        json.dumps(preparation, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(preparation, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
