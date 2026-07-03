import csv
import io
import json
import os
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from oms_v1.cli import main
from oms_v1.excel_importer import ExcelOMSImporter


class ExcelImporterTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.live_root = self.root / "live"
        self.operating_root = self.root / "operating"

    def tearDown(self):
        self.tmp.cleanup()
        for key in ["FEISHU_USER_ID_JUNE", "FEISHU_USER_ID_LIUJIE", "FEISHU_USER_ID_HUANHUAN", "FEISHU_USER_ID_NANA"]:
            os.environ.pop(key, None)

    def _csv(self, name, rows):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)
        return path

    def test_excel_sources_convert_to_user_centric_work_items(self):
        os.environ["FEISHU_USER_ID_JUNE"] = "ou_june"
        os.environ["FEISHU_USER_ID_HUANHUAN"] = "ou_huanhuan"
        os.environ["FEISHU_USER_ID_NANA"] = "ou_nana"
        resident = self._csv("resident.csv", [{"客户姓名": "张三", "房间": "301", "服务": "特殊餐和产护护理"}])
        room_status = self._csv("room.csv", [{"房号": "301", "房态": "待排房", "备注": "需要清理"}])
        contracts = self._csv("contracts.csv", [{"签约客户": "李四", "合同编号": "HJ-001", "合同金额": "10000"}])

        stream = ExcelOMSImporter(self.live_root, self.operating_root).import_sources(
            resident=resident,
            room_status=room_status,
            contracts=contracts,
        )
        assignments = {item["action_type"]: item["excel_record"]["assignment"] for item in stream["work_items"]}

        self.assertEqual(stream["record_count"], 3)
        self.assertEqual(assignments["excel_room_status_task"]["workspace"], "店总工作台")
        self.assertEqual(assignments["excel_room_status_task"]["role"], "店长 + 销售")
        self.assertEqual(assignments["excel_room_status_task"]["user_id"], "ou_june")
        self.assertEqual(assignments["excel_contract_customer_task"]["workspace"], "销售工作台")
        self.assertEqual(assignments["excel_contract_customer_task"]["user_id"], "ou_huanhuan")
        self.assertEqual(assignments["excel_resident_service_task"]["workspace"], "管家工作台")
        self.assertEqual(assignments["excel_resident_service_task"]["user_id"], "ou_nana")
        self.assertTrue((self.live_root / "pending_outbox" / "Excel_OMS导入.jsonl").exists())
        self.assertTrue((self.operating_root / "excel_work_items.jsonl").exists())

    def test_missing_user_id_stays_unresolved_and_pending(self):
        room_status = self._csv("room.csv", [{"房号": "302", "房态": "空房"}])

        stream = ExcelOMSImporter(self.live_root, self.operating_root).import_sources(room_status=room_status)
        work_item = stream["work_items"][0]

        self.assertEqual(work_item["excel_record"]["assignment"]["user_id_status"], "unresolved_user_id")
        self.assertEqual(work_item["status"], "attention_required")
        self.assertTrue(work_item["confirmation_required"])
        self.assertEqual(stream["pending_outbox_count"], 1)

    def test_xlsx_source_is_supported(self):
        from openpyxl import Workbook

        path = self.root / "resident.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户姓名", "房间", "服务"])
        sheet.append(["赵六", "502", "产护护理"])
        workbook.save(path)

        stream = ExcelOMSImporter(self.live_root, self.operating_root).import_sources(resident=path)

        self.assertEqual(stream["records"][0]["normalized"]["customer_name"], "赵六")
        self.assertEqual(stream["work_items"][0]["workspace"], "管家工作台")

    def test_xlsx_imports_all_business_sheets_and_skips_non_business_sheets(self):
        from openpyxl import Workbook

        path = self.root / "contracts_multi_sheet.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2026.07"
        sheet.append(["凰家母婴空间签约客户一览表", None, None])
        sheet.append(["序号", "签约日期", "姓名", "价格"])
        sheet.append([1, "2026.7.1", "客户A", 20000])
        next_sheet = workbook.create_sheet("2026.08")
        next_sheet.append(["凰家母婴空间签约客户一览表", None, None])
        next_sheet.append(["序号", "签约日期", "姓名", "价格"])
        next_sheet.append([1, "2026.8.1", "客户B", 30000])
        password_sheet = workbook.create_sheet("楼层密码")
        password_sheet.append(["无线", "密码"])
        password_sheet.append(["hjmy", "88888888"])
        workbook.save(path)

        stream = ExcelOMSImporter(self.live_root, self.operating_root).import_sources(contracts=path)

        self.assertEqual(stream["record_count"], 2)
        self.assertEqual({record["source_sheet"] for record in stream["records"]}, {"2026.07", "2026.08"})
        self.assertEqual([record["row_number"] for record in stream["records"]], [3, 3])
        self.assertEqual({item["daily_process"] for item in stream["work_items"]}, {"签约客户提报"})

    def test_xlsx_skips_placeholder_rows_without_business_values(self):
        from openpyxl import Workbook

        path = self.root / "contracts_placeholder_rows.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["序号", "签约日期", "姓名", "价格"])
        sheet.append([1, "", "", ""])
        sheet.append([2, "2026.7.1", "客户A", 20000])
        workbook.save(path)

        stream = ExcelOMSImporter(self.live_root, self.operating_root).import_sources(contracts=path)

        self.assertEqual(stream["record_count"], 1)
        self.assertEqual(stream["records"][0]["normalized"]["customer_name"], "客户A")
        self.assertEqual(stream["records"][0]["normalized"]["amount"], "20000")

    def test_cli_excel_import_outputs_json(self):
        contracts = self._csv("contracts.csv", [{"签约客户": "王五", "合同编号": "HJ-002"}])
        output = io.StringIO()
        with redirect_stdout(output):
            code = main(
                [
                    "excel-import",
                    "--contracts",
                    str(contracts),
                    "--live-root",
                    str(self.live_root),
                    "--operating-root",
                    str(self.operating_root),
                    "--pretty",
                ]
            )
        payload = json.loads(output.getvalue())

        self.assertEqual(code, 0)
        self.assertEqual(payload["schema_version"], "oms.v1.excel_import_stream")
        self.assertEqual(payload["work_items"][0]["workspace"], "销售工作台")


if __name__ == "__main__":
    unittest.main()
