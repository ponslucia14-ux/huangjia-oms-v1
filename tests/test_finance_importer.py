import csv
import io
import json
import os
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from oms_v1.cli import main
from oms_v1.finance_importer import FinanceDataImporter


class FinanceImporterTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.live_root = self.root / "live"
        self.operating_root = self.root / "operating"

    def tearDown(self):
        self.tmp.cleanup()
        for key in ["FEISHU_USER_ID_LIUJIE", "FEISHU_USER_ID_HUANHUAN", "FEISHU_USER_ID_SHILEI"]:
            os.environ.pop(key, None)

    def _csv(self, name, rows):
        path = self.root / name
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)
        return path

    def _realworld_mapping(self, rows):
        path = self.live_root / "realworld_mapping" / "OMS_RealWorld_Mapping.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps({"rows": rows}, ensure_ascii=False), encoding="utf-8")
        return path

    def test_finance_sources_generate_events_settlements_and_work_items(self):
        self._realworld_mapping(
            [
                {"name": "刘晶", "role": "财务", "user_id": "ou_liujie"},
                {"name": "杨欢欢", "role": "销售", "user_id": "ou_huanhuan"},
                {"name": "石磊", "role": "boss", "user_id": "ou_boss"},
            ]
        )
        finance_daily = self._csv("daily.csv", [{"日期": "2026.7.1", "收入项目": "入住尾款", "收入金额": "10000"}])
        commission = self._csv("commission.csv", [{"宝妈姓名": "张三", "销售": "杨欢欢", "提成": "900"}])
        wage = self._csv("wage.csv", [{"姓名": "照护师A", "应发": "6000", "备注": "5月工资"}])

        stream = FinanceDataImporter(self.live_root, self.operating_root).import_sources(
            finance_daily=finance_daily,
            sales_commission=commission,
            care_wage=wage,
        )
        assignments = {item["action_type"]: item["finance_record"]["assignment"] for item in stream["work_items"]}

        self.assertEqual(stream["record_count"], 3)
        self.assertEqual(stream["financial_event_count"], 3)
        self.assertEqual(stream["settlement_record_count"], 3)
        self.assertEqual(stream["pending_outbox_count"], 3)
        self.assertEqual(assignments["finance_daily_report_task"]["role"], "财务")
        self.assertEqual(assignments["finance_daily_report_task"]["workspace"], "财务工作台")
        self.assertEqual(assignments["finance_daily_report_task"]["user_id"], "ou_liujie")
        self.assertEqual(assignments["finance_sales_commission_task"]["role"], "销售")
        self.assertEqual(assignments["finance_sales_commission_task"]["user_id"], "ou_huanhuan")
        self.assertEqual(assignments["finance_care_wage_task"]["role"], "管理")
        self.assertEqual(assignments["finance_care_wage_task"]["user_id"], "ou_boss")
        evidence = stream["records"][0]["source_evidence"]
        self.assertEqual(evidence["truth_source"], "Finance Excel")
        self.assertEqual(evidence["source_file"], str(finance_daily))
        self.assertEqual(evidence["row_number"], 2)
        self.assertEqual(stream["financial_events"][0]["truth_status"], "source_verified")
        self.assertEqual(stream["settlement_records"][0]["source_evidence"]["record_id"], stream["records"][0]["record_id"])
        self.assertTrue((self.operating_root / "finance_work_items.jsonl").exists())
        self.assertTrue((self.live_root / "finance" / "financial_events.jsonl").exists())
        self.assertTrue((self.live_root / "finance" / "settlement_records.jsonl").exists())
        self.assertTrue((self.live_root / "pending_outbox" / "Finance_OMS导入.jsonl").exists())

    def test_missing_user_id_is_required_and_pending(self):
        finance_daily = self._csv("daily.csv", [{"日期": "2026.7.1", "支出项目": "采购", "支出金额": "300"}])

        stream = FinanceDataImporter(self.live_root, self.operating_root).import_sources(finance_daily=finance_daily)
        work_item = stream["work_items"][0]

        self.assertEqual(work_item["finance_record"]["assignment"]["user_id_status"], "missing_required_user_id")
        self.assertEqual(work_item["status"], "attention_required")
        self.assertTrue(work_item["confirmation_required"])

    def test_xlsx_multi_sheet_finance_source_is_supported(self):
        from openpyxl import Workbook

        path = self.root / "finance_daily.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "日报"
        sheet.append(["凰家母婴空间财务日报表", None, None])
        sheet.append(["日期", "收入项目", "收入金额"])
        sheet.append(["2026.7.1", "入住尾款", 10000])
        empty = workbook.create_sheet("空白")
        empty.append(["说明"])
        workbook.save(path)

        stream = FinanceDataImporter(self.live_root, self.operating_root).import_sources(finance_daily=path)

        self.assertEqual(stream["record_count"], 1)
        self.assertEqual(stream["records"][0]["source_sheet"], "日报")
        self.assertEqual(stream["financial_events"][0]["income_amount"], "10000")

    def test_real_income_crosstab_is_split_by_month(self):
        from openpyxl import Workbook

        path = self.root / "real_income.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "实入账"
        sheet.append(["凰 家 母 婴 空 间 实 入 账"])
        sheet.append([])
        sheet.append([None, "1月份", "2月份"])
        sheet.append(["内店实际入住金额（服务金额：元）", 1000, 2000])
        workbook.save(path)

        stream = FinanceDataImporter(self.live_root, self.operating_root).import_sources(real_income=path)

        self.assertEqual(stream["record_count"], 2)
        self.assertEqual([record["normalized"]["date"] for record in stream["records"]], ["1月份", "2月份"])
        self.assertEqual([event["income_amount"] for event in stream["financial_events"]], ["1000", "2000"])

    def test_cli_finance_import_outputs_json(self):
        finance_daily = self._csv("daily.csv", [{"日期": "2026.7.1", "收入项目": "入住尾款", "收入金额": "10000"}])
        output = io.StringIO()
        with redirect_stdout(output):
            code = main(
                [
                    "finance-import",
                    "--finance-daily",
                    str(finance_daily),
                    "--live-root",
                    str(self.live_root),
                    "--operating-root",
                    str(self.operating_root),
                    "--pretty",
                ]
            )
        payload = json.loads(output.getvalue())

        self.assertEqual(code, 0)
        self.assertEqual(payload["schema_version"], "oms.v1.finance_import_stream")
        self.assertEqual(payload["work_items"][0]["workspace"], "财务工作台")


if __name__ == "__main__":
    unittest.main()
